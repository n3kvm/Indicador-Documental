import json
import os
import math
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl


TEXT_SUMMARY = Path(os.environ.get("PDF_TEXT_SUMMARY_PATH", "work/pdf_text_summary.json"))
CRON = Path(os.environ.get("CRONOGRAMA_PATH", "work/cronograma_mayo_2026_copia.xlsx"))
COMPARE = Path(os.environ.get("COMPARE_OUTPUT_PATH", r"C:\Users\DELL\Documents\Codex\2026-06-03\files-mentioned-by-the-user-05\outputs\validacion_soportes_sharepoint_mayo_2026.xlsx"))
OUT = Path(os.environ.get("PILLARS_VALIDATION_PATH", "work/pillars_hours_validation.json"))
MONTH_NAME = os.environ.get("SHAREPOINT_MES_NOMBRE", "").upper()

PILLAR_ALIASES = {
    "INS-ELE": "INS-ELE",
    "SEÑA": "SEÑA",
    "SENA": "SEÑA",
    "SEN-MTTO": "SEÑA",
    "OBR-CIV": "OBR-CIV",
    "CARP-MET": "CARP-MET",
    "CHA-LLAV": "CHA-LLAV",
    "PINT MTTO": "PINT MTTO",
    "PINT": "PINT MTTO",
    "HID-SAN": "HID-SAN",
    "CUBI": "CUBI",
    "CUB-MTTO": "CUBI",
    "APY-LOG": "APY-LOG",
    "APY-MTTO": "APY-MTTO",
    "CARP-MAD": "CARP-MAD",
    "ADE-OFI": "ADE-OFI",
    "AD-OFIC": "ADE-OFI",
    "PERSIA": "PERSIA",
    "PERS": "PERSIA",
}

PILLAR_COLS = {
    "HID-SAN": 19,
    "INS-ELE": 21,
    "OBR-CIV": 23,
    "APY-LOG": 25,
    "APY-MTTO": 27,
    "CARP-MET": 29,
    "CHA-LLAV": 31,
    "CARP-MAD": 33,
    "PINT MTTO": 35,
    "ADE-OFI": 37,
    "CUBI": 39,
    "SEÑA": 41,
    "PERSIA": 43,
}

PILLAR_HEADER_NAMES = {
    "HID-SAN": ["HIDRO. SANIT", "HID-SAN"],
    "INS-ELE": ["ELECT", "INS-ELE"],
    "OBR-CIV": ["O. CIVIL", "OBR-CIV"],
    "APY-LOG": ["APY-LOG"],
    "APY-MTTO": ["APY-MTTO"],
    "CARP-MET": ["CARP. MET", "CARP-MET"],
    "CHA-LLAV": ["CHAP-LLAV", "CHA-LLAV"],
    "CARP-MAD": ["CARP. MAD", "CARP-MAD"],
    "PINT MTTO": ["PINT", "PINT MTTO"],
    "ADE-OFI": ["AD. OFIC", "ADE-OFI"],
    "CUBI": ["CUBIERT", "CUBI"],
    "SEÑA": ["SEÑAL", "SENAL", "SEÑA"],
    "PERSIA": ["PERS", "PERSIA"],
}


def fix_text(text):
    replacements = {
        "Ã³": "ó", "Ã©": "é", "Ã¡": "á", "Ã­": "í", "Ãº": "ú", "Ã±": "ñ",
        "Ã“": "Ó", "Ã‰": "É", "Ã": "Á", "Ã": "Í", "Ãš": "Ú", "Ã‘": "Ñ",
        "ï¬": "fi", "ï¬‚": "fl", "ﬁ": "fi", "ﬂ": "fl", "Â": "",
    }
    for bad, good in replacements.items():
        text = text.replace(bad, good)
    return text


def normalize(text):
    text = fix_text(str(text or ""))
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^A-Za-z0-9]+", " ", text).lower().strip()
    return re.sub(r"\s+", " ", text)


def header_norm(value):
    return normalize(value).upper()


def find_row_with_headers(sheet, required, max_rows=12):
    required = [header_norm(r) for r in required]
    for row in range(1, min(sheet.max_row, max_rows) + 1):
        values = [header_norm(sheet.cell(row, col).value) for col in range(1, sheet.max_column + 1)]
        joined = " | ".join(values)
        if all(req in joined for req in required):
            return row
    return None


def header_map(sheet, row):
    out = {}
    for col in range(1, sheet.max_column + 1):
        key = header_norm(sheet.cell(row, col).value)
        if key:
            out[key] = col
    return out


def find_col(headers, *names):
    normalized = [header_norm(name) for name in names]
    for name in normalized:
        if name in headers:
            return headers[name]
    for key, col in headers.items():
        if any(name in key for name in normalized):
            return col
    return None


def find_cronog_sheet(workbook):
    candidates = []
    for sheet in workbook.worksheets:
        header_row = find_row_with_headers(sheet, ["SEDES PROGRAMADAS", "FECHA REAL DE EJECUCIÓN", "TOTAL AUTOGESTIONES"])
        if header_row:
            score = 20 if "CRONOG" in header_norm(sheet.title) else 0
            if MONTH_NAME and MONTH_NAME in header_norm(sheet.title):
                score += 100
            candidates.append((score, sheet, header_row))
    if not candidates:
        raise KeyError(f"No encontré hoja Cronog con fechas reales. Hojas disponibles: {workbook.sheetnames}")
    if MONTH_NAME and not any(MONTH_NAME in header_norm(item[1].title) for item in candidates):
        raise KeyError(f"No encontré hoja Cronog para {MONTH_NAME}. Hojas disponibles: {workbook.sheetnames}")
    return sorted(candidates, key=lambda item: item[0], reverse=True)[0]


def score_name(a, b):
    na, nb = normalize(a), normalize(b)
    if not na or not nb:
        return 0
    if na == nb:
        return 1
    seq = SequenceMatcher(None, na, nb).ratio()
    ta, tb = set(na.split()), set(nb.split())
    token = len(ta & tb) / max(len(ta | tb), 1)
    containment = len(ta & tb) / max(min(len(ta), len(tb)), 1)
    return max(seq, token, containment * 0.92)


def more_specific_site(candidate, current):
    if not current:
        return True
    return len(normalize(candidate).split()) > len(normalize(current).split())


def parse_time(value):
    m = re.search(r"(\d{1,2}):(\d{2})", value or "")
    if not m:
        return None
    return int(m.group(1)) * 60 + int(m.group(2))


def hours_between(start, end):
    if start is None or end is None:
        return None
    if end < start:
        end += 24 * 60
    return (end - start) / 60


def effective_hours(start, end):
    gross = hours_between(start, end)
    if gross is None:
        return None
    lunch = 1 if gross >= 6 else 0
    return max(gross - lunch, 0)


def parse_date(text):
    m = re.search(r"Fecha de realizaci[oó]n\s+(\d{1,2})-([A-Za-z]{3})-(\d{4})", text, re.I)
    if not m:
        return ""
    months = {"jan": 1, "ene": 1, "feb": 2, "mar": 3, "apr": 4, "abr": 4, "may": 5, "jun": 6, "jul": 7, "aug": 8, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12, "dic": 12}
    month = months.get(m.group(2).lower()[:3], 0)
    if not month:
        return ""
    return f"{int(m.group(3)):04d}-{month:02d}-{int(m.group(1)):02d}"


def extract_technicians(block):
    m = re.search(r"Nombre de t[eé]cnico\s*(.*?)(?:LISTA DE CHEQUEO|Item Zona|ACTIVIDADES REALIZADAS)", block, re.S | re.I)
    if not m:
        return []
    lines = [fix_text(x).strip() for x in m.group(1).splitlines()]
    ignore = {"", "in-house", "movil", "móvil"}
    techs = []
    for line in lines:
        if normalize(line) in ignore:
            continue
        if len(line) >= 5 and not re.search(r"^\d+$", line):
            # Some PDFs wrap the last surname onto a separate line, e.g.
            # "CRISTIAN ANTONIO CASTILLO" / "COMETA". Treat a lone word as a
            # continuation of the previous name instead of another technician.
            if techs and len(line.split()) == 1 and line.isupper():
                techs[-1] = f"{techs[-1]} {line}"
            else:
                techs.append(line)
    return techs


def extract_activity_section(block):
    m = re.search(r"ACTIVIDADES REALIZADAS(.*?)(?:INSUMOS UTILIZADOS|Observaciones Generales|T[eé]cnico Brillaseo|Aprobado usuario|$)", block, re.S | re.I)
    return m.group(1) if m else ""


def extract_pillars(activity_text):
    counts = Counter()
    for raw in re.findall(r"\(([A-ZÑÁÉÍÓÚÜ0-9 -]{3,})\)", fix_text(activity_text)):
        code = re.sub(r"\s+", " ", raw.strip().upper())
        code = code.replace("AD OFIC", "ADE-OFI")
        canonical = PILLAR_ALIASES.get(code)
        if canonical:
            counts[canonical] += 1
    return counts


def extract_charge_counts(block, activity_text, pillars):
    total_match = re.search(r"Total autogestiones\s+(\d+)", fix_text(block), re.I)
    total = int(total_match.group(1)) if total_match else sum(pillars.values())
    if total <= 0:
        return 0, 0, 0

    material_nums = set()
    insumos = re.search(
        r"INSUMOS UTILIZADOS(.*?)(?:Observaciones Generales|T[eÃ©]cnico Brillaseo|Aprobado usuario|$)",
        fix_text(block),
        re.S | re.I,
    )
    if insumos:
        for raw in re.findall(r"(?m)^\s*(\d{1,3})\s+\S", insumos.group(1)):
            num = int(raw)
            if 1 <= num <= total:
                material_nums.add(num)

    text = fix_text(activity_text)
    starts = []
    for match in re.finditer(r"(?m)^\s*(\d{1,3})(?:\s|$)", text):
        num = int(match.group(1))
        if 1 <= num <= total:
            starts.append((num, match.start()))
    starts.sort(key=lambda item: item[1])
    for idx, (num, start) in enumerate(starts):
        end = starts[idx + 1][1] if idx + 1 < len(starts) else len(text)
        segment = text[start:end]
        if re.search(r"(?im)^\s*Materiales\s*$", segment):
            material_nums.add(num)

    con_cobro = len(material_nums)
    sin_cobro = max(total - con_cobro, 0)
    return total, con_cobro, sin_cobro


def split_checklists(text):
    parts = re.split(r"(?=No\. Autogestion\s*\n?\s*\d+)", text)
    return [p for p in parts if re.search(r"No\. Autogestion", p)]


def read_cron_rows():
    wb = openpyxl.load_workbook(CRON, data_only=True)
    _, ws, header_row = find_cronog_sheet(wb)
    headers = header_map(ws, header_row)
    col_sede = find_col(headers, "SEDES PROGRAMADAS")
    col_horas = find_col(headers, "HORAS REALES EJ")
    col_autog = find_col(headers, "TOTAL AUTOGESTIONES")
    pillar_cols = {}
    for code, names in PILLAR_HEADER_NAMES.items():
        pillar_cols[code] = find_col(headers, *names) or PILLAR_COLS.get(code)
    rows = {}
    for r in range(header_row + 1, ws.max_row + 1):
        sede = ws.cell(r, col_sede).value
        if not sede or str(sede).upper().startswith("ZONA"):
            continue
        row = {
            "sede": str(sede).strip(),
            "horas_reales_cronograma": ws.cell(r, col_horas).value or 0 if col_horas else 0,
            "total_autogestiones_cronograma": ws.cell(r, col_autog).value or 0 if col_autog else 0,
        }
        for code, col in pillar_cols.items():
            row[f"cron_{code}"] = ws.cell(r, col).value or 0
        rows[normalize(sede)] = row
    return rows


def read_support_completeness():
    wb = openpyxl.load_workbook(COMPARE, data_only=True)
    ws = wb["Resumen por sede"]
    headers = [cell.value for cell in ws[1]]
    out = {}
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        out[normalize(row.get("sede"))] = row
    return out


def read_file_site_matches():
    wb = openpyxl.load_workbook(COMPARE, data_only=True)
    if "Archivos SharePoint" not in wb.sheetnames:
        return {}
    ws = wb["Archivos SharePoint"]
    headers = [cell.value for cell in ws[1]]
    out = {}
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        archivo = row.get("archivo")
        sede_match = row.get("sede_match")
        if archivo and sede_match:
            try:
                confidence = float(row.get("confianza_match") or 0)
            except (TypeError, ValueError):
                confidence = 0
            out[str(archivo)] = {
                "sede": str(sede_match).strip(),
                "confianza": confidence,
                "sede_archivo_limpia": str(row.get("sede_archivo") or "").strip(),
            }
    return out


summary = json.loads(TEXT_SUMMARY.read_text(encoding="utf-8"))
cron_rows = read_cron_rows()
support_rows = read_support_completeness()
file_site_matches = read_file_site_matches()
canonical_sites = {}
for key, row in cron_rows.items():
    canonical_sites[key] = row["sede"]
for key, row in support_rows.items():
    canonical_sites.setdefault(key, row.get("sede", ""))


def canonical_key(site):
    if not site:
        return ""
    exact_key = normalize(site)
    if exact_key in canonical_sites:
        return exact_key
    best_key, best_score = normalize(site), 0
    for key, name in canonical_sites.items():
        s = score_name(site, name)
        if s > best_score or (abs(s - best_score) < 0.0001 and more_specific_site(name, canonical_sites.get(best_key, ""))):
            best_key, best_score = key, s
    return best_key if best_score >= 0.62 else normalize(site)

checklists = []
for item in summary:
    if not item.get("text_path"):
        continue
    text = fix_text(Path(item["text_path"]).read_text(encoding="utf-8", errors="ignore"))
    for block in split_checklists(text):
        auto_match = re.search(r"No\. Autogestion\s*(\d+)", block, re.I)
        sede = ""
        m = re.search(r"Sede \(Comfandi\)\s*(.*?)(?:Centro de costo|Fecha de realizaci[oó]n)", block, re.S | re.I)
        if m:
            sede = " ".join(x.strip() for x in fix_text(m.group(1)).splitlines() if x.strip())
        date = parse_date(block)
        start = parse_time(re.search(r"Hora inicial\s+([0-9:]+)", block, re.I).group(1)) if re.search(r"Hora inicial\s+([0-9:]+)", block, re.I) else None
        end_match = re.search(r"Hora\s+(?:final|finaI|fi?nal)\s+([0-9:]+)", fix_text(block), re.I)
        end = parse_time(end_match.group(1)) if end_match else None
        total_match = re.search(r"Total horas\s+([0-9:]+)", block, re.I)
        total_pdf = hours_between(0, parse_time(total_match.group(1))) if total_match else None
        techs = extract_technicians(block)
        eff = effective_hours(start, end)
        activity_section = extract_activity_section(block)
        pillars = extract_pillars(activity_section)
        actividades_total, con_cobro, sin_cobro = extract_charge_counts(block, activity_section, pillars)
        file_match = file_site_matches.get(item["Name"], {})
        file_site = file_match.get("sede", "") if isinstance(file_match, dict) else str(file_match or "")
        file_confidence = file_match.get("confianza", 0) if isinstance(file_match, dict) else 0
        file_site_clean = file_match.get("sede_archivo_limpia", "") if isinstance(file_match, dict) else ""
        sede_final = file_site or sede
        sede_key = canonical_key(sede_final)
        pdf_key = canonical_key(sede)
        file_key = canonical_key(file_site)
        sede_mismatch = bool(file_confidence >= 0.9 and file_site and sede and pdf_key and file_key and pdf_key != file_key)
        checklists.append({
            "autogestion": auto_match.group(1) if auto_match else "",
            "archivo": item["Name"],
            "sede_pdf": sede_final,
            "sede_pdf_original": sede,
            "sede_archivo": file_site,
            "sede_archivo_limpia": file_site_clean,
            "confianza_sede_archivo": file_confidence,
            "sede_key": sede_key,
            "sede_mismatch_archivo_pdf": sede_mismatch,
            "fecha": date,
            "hora_inicial": f"{start//60:02d}:{start%60:02d}" if start is not None else "",
            "hora_final": f"{end//60:02d}:{end%60:02d}" if end is not None else "",
            "horas_pdf_gross": total_pdf,
            "horas_efectivas_por_tecnico": eff,
            "tecnicos": len(techs),
            "tecnicos_nombres": "; ".join(techs),
            "horas_calculadas": None if eff is None else eff * max(len(techs), 1),
            "autogestiones_pdf": sum(pillars.values()),
            "actividades_total_pdf": actividades_total,
            "actividades_con_cobro_pdf": con_cobro,
            "actividades_sin_cobro_pdf": sin_cobro,
            "pilares": dict(pillars),
        })

by_site = {}
for chk in checklists:
    key = chk["sede_key"]
    if not key:
        key = normalize(chk["archivo"])
    site = by_site.setdefault(key, {
        "sede": canonical_sites.get(key, chk["sede_pdf"]),
        "checklists_pdf": 0,
        "horas_calculadas_pdf": 0,
        "autogestiones_pdf": 0,
        "actividades_total_pdf": 0,
        "actividades_con_cobro_pdf": 0,
        "actividades_sin_cobro_pdf": 0,
        "pilares_pdf": Counter(),
        "fechas_pdf": set(),
        "archivos": set(),
    })
    site["checklists_pdf"] += 1
    site["horas_calculadas_pdf"] += chk["horas_calculadas"] or 0
    site["autogestiones_pdf"] += chk["autogestiones_pdf"]
    site["actividades_total_pdf"] += chk["actividades_total_pdf"]
    site["actividades_con_cobro_pdf"] += chk["actividades_con_cobro_pdf"]
    site["actividades_sin_cobro_pdf"] += chk["actividades_sin_cobro_pdf"]
    site["pilares_pdf"].update(chk["pilares"])
    if chk["fecha"]:
        site["fechas_pdf"].add(chk["fecha"])
    site["archivos"].add(chk["archivo"])

site_rows = []
for key in sorted(set(by_site) | set(support_rows)):
    row = by_site.get(key, {
        "sede": support_rows.get(key, {}).get("sede", canonical_sites.get(key, key)),
        "checklists_pdf": 0,
        "horas_calculadas_pdf": 0,
        "autogestiones_pdf": 0,
        "actividades_total_pdf": 0,
        "actividades_con_cobro_pdf": 0,
        "actividades_sin_cobro_pdf": 0,
        "pilares_pdf": Counter(),
        "fechas_pdf": set(),
        "archivos": set(),
    })
    cron = cron_rows.get(key, {})
    support = support_rows.get(key, {})
    out = {
        "sede": row["sede"],
        "estado_soportes": support.get("estado", ""),
        "cantidad_digitada": support.get("cantidad_digitada", ""),
        "cantidad_cubierta": support.get("cantidad_cubierta", ""),
        "observacion_soportes": support.get("observacion", ""),
        "checklists_pdf": row["checklists_pdf"],
        "fechas_pdf": ",".join(sorted(row["fechas_pdf"])),
        "archivos": " | ".join(sorted(row["archivos"])),
        "horas_calculadas_pdf": round(row["horas_calculadas_pdf"], 2),
        "horas_reales_cronograma": cron.get("horas_reales_cronograma", ""),
        "dif_horas_pdf_vs_cronograma": "" if not cron else round(row["horas_calculadas_pdf"] - float(cron.get("horas_reales_cronograma") or 0), 2),
        "autogestiones_pdf": row["autogestiones_pdf"],
        "actividades_total_pdf": row["actividades_total_pdf"],
        "actividades_con_cobro_pdf": row["actividades_con_cobro_pdf"],
        "actividades_sin_cobro_pdf": row["actividades_sin_cobro_pdf"],
        "total_autogestiones_cronograma": cron.get("total_autogestiones_cronograma", ""),
        "dif_autogestiones_pdf_vs_cronograma": "" if not cron else row["autogestiones_pdf"] - float(cron.get("total_autogestiones_cronograma") or 0),
    }
    for code in sorted(PILLAR_COLS):
        out[f"pdf_{code}"] = row["pilares_pdf"].get(code, 0)
        out[f"cron_{code}"] = cron.get(f"cron_{code}", "")
        out[f"dif_{code}"] = "" if not cron else row["pilares_pdf"].get(code, 0) - float(cron.get(f"cron_{code}") or 0)
    site_rows.append(out)

alerts = []
for item in summary:
    if item.get("downloadedBytes") == 0 or int(item.get("Length") or 0) == 0:
        alerts.append({
            "categoria": "Archivo",
            "severidad": "Crítica",
            "tipo": "PDF vacío o sin contenido",
            "sede_pdf": "",
            "autogestion": "",
            "fecha": "",
            "campo": "Tamaño archivo",
            "valor_pdf": item.get("downloadedBytes", 0),
            "valor_cronograma": "",
            "diferencia": "",
            "actividades_total_pdf": "",
            "actividades_con_cobro_pdf": "",
            "actividades_sin_cobro_pdf": "",
            "detalle": f"SharePoint reporta el archivo con {item.get('Length', 0)} bytes y la descarga local quedó en {item.get('downloadedBytes', 0)} bytes. No se puede leer fecha, checklist, horas ni pilares.",
            "archivo": item.get("Name", ""),
        })

for chk in checklists:
    if chk.get("sede_mismatch_archivo_pdf"):
        alerts.append({
            "categoria": "Lectura PDF",
            "severidad": "Revisar",
            "tipo": "Sede del PDF diferente a sede asociada",
            "detalle": f"La sede asociada por el nombre limpio del archivo es {chk.get('sede_archivo')}, pero dentro del PDF aparece {chk.get('sede_pdf_original')}. El nombre limpio usado fue: {chk.get('sede_archivo_limpia')}. Se ignoran observaciones como pendiente firma, sin índice o por firma de gestor.",
            "campo": "Sede (Comfandi)",
            "valor_pdf": chk.get("sede_pdf_original", ""),
            "valor_cronograma": chk.get("sede_archivo", ""),
            **chk,
        })
    if not chk["fecha"] or not chk["hora_inicial"] or not chk["hora_final"]:
        alerts.append({"categoria": "Horas", "severidad": "Revisar", "tipo": "Dato horario incompleto", **chk})
    if chk["tecnicos"] == 0:
        alerts.append({"categoria": "Lectura PDF", "severidad": "Revisar", "tipo": "Sin técnico detectado", **chk})
    if chk["autogestiones_pdf"] == 0:
        alerts.append({"categoria": "Pilares", "severidad": "Revisar", "tipo": "Sin pilares detectados en actividades", **chk})
    if chk["horas_efectivas_por_tecnico"] is not None and chk["horas_efectivas_por_tecnico"] > 8.5:
        alerts.append({"categoria": "Horas", "severidad": "Revisar", "tipo": "Jornada efectiva superior a 8 horas", **chk})

site_alerts = []
for row in site_rows:
    estado = row.get("estado_soportes") or "Sin cruce"
    support_complete = estado in {"Completo", "Completo con observación"}
    missing_support = int(float(row.get("cantidad_digitada") or 0)) - int(float(row.get("cantidad_cubierta") or 0)) if row.get("cantidad_digitada") != "" else 0
    dif_aut = row.get("dif_autogestiones_pdf_vs_cronograma")
    dif_hours = row.get("dif_horas_pdf_vs_cronograma")

    if estado == "Completo con observación":
        site_alerts.append({
            "categoria": "Soportes",
            "severidad": "Revisar",
            "tipo": "Completo con fecha diferente",
            "sede": row["sede"],
            "detalle": row.get("observacion_soportes") or "La cantidad de soportes está completa, pero las fechas del PDF no coinciden exactamente con el cronograma.",
            "valor_pdf": row.get("cantidad_cubierta", ""),
            "valor_cronograma": row.get("cantidad_digitada", ""),
            "diferencia": 0,
            "campo": "Fechas PDF vs cronograma",
        })

    if estado == "Incompleto":
        detail = f"Faltan {missing_support} checklist(s) cargados para poder cerrar la validación mensual."
        site_alerts.append({
            "categoria": "Soportes",
            "severidad": "Crítica",
            "tipo": "Soportes faltantes",
            "sede": row["sede"],
            "detalle": detail,
            "valor_pdf": row.get("cantidad_cubierta", ""),
            "valor_cronograma": row.get("cantidad_digitada", ""),
            "diferencia": -missing_support,
            "campo": "Checklists cargados vs digitados",
        })

    if dif_aut != "" and abs(float(dif_aut)) > 0.01:
        site_alerts.append({
            "categoria": "Autogestiones",
            "severidad": "Crítica" if support_complete else "Informativa",
            "tipo": "Diferencia de autogestiones",
            "sede": row["sede"],
            "detalle": (
                "Los soportes completos no coinciden con el total digitado en cronograma."
                if support_complete
                else "La diferencia existe, pero la sede tiene soportes incompletos; no es concluyente hasta cargar todos los PDFs."
            ),
            "valor_pdf": row.get("autogestiones_pdf", ""),
            "valor_cronograma": row.get("total_autogestiones_cronograma", ""),
            "diferencia": dif_aut,
            "campo": "TOTAL AUTOGESTIONES",
        })

    if dif_hours != "" and abs(float(dif_hours)) > 0.25:
        site_alerts.append({
            "categoria": "Horas",
            "severidad": "Crítica" if support_complete else "Informativa",
            "tipo": "Diferencia de horas",
            "sede": row["sede"],
            "detalle": (
                "Las horas calculadas desde PDFs completos no coinciden con horas reales del cronograma."
                if support_complete
                else "La diferencia de horas no es concluyente porque faltan soportes."
            ),
            "valor_pdf": row.get("horas_calculadas_pdf", ""),
            "valor_cronograma": row.get("horas_reales_cronograma", ""),
            "diferencia": dif_hours,
            "campo": "HORAS REALES EJ",
        })

    if support_complete:
        for code in sorted(PILLAR_COLS):
            dif = row.get(f"dif_{code}")
            if dif != "" and abs(float(dif)) > 0.01:
                site_alerts.append({
                    "categoria": "Pilares",
                    "severidad": "Revisar",
                    "tipo": f"Diferencia pilar {code}",
                    "sede": row["sede"],
                    "detalle": "El conteo del pilar detectado en PDFs completos no coincide con el cronograma.",
                    "valor_pdf": row.get(f"pdf_{code}", ""),
                    "valor_cronograma": row.get(f"cron_{code}", ""),
                    "diferencia": dif,
                    "campo": code,
                })

OUT.write_text(json.dumps({
    "checklists": checklists,
    "site_rows": site_rows,
    "alerts": alerts,
    "site_alerts": site_alerts,
    "summary": {
        "pdfs_leidos": len({c["archivo"] for c in checklists}),
        "checklists_leidos": len(checklists),
        "horas_calculadas": round(sum(c["horas_calculadas"] or 0 for c in checklists), 2),
        "autogestiones_detectadas": sum(c["autogestiones_pdf"] for c in checklists),
        "alertas_checklist": len(alerts),
        "alertas_sede": len(site_alerts),
        "alertas": len(alerts) + len(site_alerts),
    },
}, ensure_ascii=False, indent=2), encoding="utf-8")
print(OUT.read_text(encoding="utf-8")[:1000])
