import json
import os
import re
import unicodedata
from calendar import monthrange
from collections import defaultdict
from datetime import date, timedelta
from difflib import SequenceMatcher
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


EXPECTED = Path(os.environ.get("EXPECTED_SCHEDULE_PATH", "work/expected_schedule.json"))
SP_FILES = Path(os.environ.get("SHAREPOINT_FILES_PATH", "work/sharepoint_files.json"))
PDF_TEXT_SUMMARY = Path(os.environ.get("PDF_TEXT_SUMMARY_PATH", "work/pdf_text_summary.json"))
OUT = Path(os.environ.get("COMPARE_OUTPUT_PATH", r"C:\Users\DELL\Documents\Codex\2026-06-03\files-mentioned-by-the-user-05\outputs\validacion_soportes_sharepoint_mayo_2026.xlsx"))
YEAR = int(os.environ.get("SHAREPOINT_ANIO", "2026"))
MONTH = int(os.environ.get("SHAREPOINT_MES", "5"))


def normalize(text):
    text = "" if text is None else str(text)
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = text.lower()
    text = re.sub(r"\.[a-z0-9]+$", "", text)
    text = re.sub(r"^\s*\d+\s*[-.]?\s*", "", text)
    text = re.sub(r"\(([^)]*)\)", r" \1 ", text)
    text = text.replace("ips ", "ips ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    stop = {"comfandi", "sede", "de", "la", "las", "los", "el", "y"}
    tokens = [t for t in text.split() if t not in stop]
    return " ".join(tokens)


def sharepoint_file_view_url(file_row):
    origin = file_row.get("SiteOrigin") or "https://comfandisa.sharepoint.com"
    server_relative = file_row.get("ServerRelativeUrl", "")
    if not server_relative:
        return file_row.get("WebUrl") or file_row.get("webUrl") or ""
    encoded_path = quote(server_relative, safe="/")
    return f"{origin}/:b:/r{encoded_path}?csf=1&web=1"


def parse_days_text(text):
    text = normalize_day_text(text)
    days = set()
    for a, b in re.findall(r"(\d{1,2})\s*(?:a|al)\s*(\d{1,2})", text):
        start, end = int(a), int(b)
        if 1 <= start <= 31 and 1 <= end <= 31:
            lo, hi = sorted((start, end))
            days.update(range(lo, hi + 1))
    consumed = re.sub(r"\d{1,2}\s*(?:a|al)\s*\d{1,2}", " ", text)
    for d in re.findall(r"\d{1,2}", consumed):
        day = int(d)
        if 1 <= day <= 31:
            days.add(day)
    return days


def normalize_day_text(text):
    text = unicodedata.normalize("NFKD", str(text)).encode("ascii", "ignore").decode("ascii").lower()
    text = text.replace("–", "-").replace("—", "-")
    return text


def fix_text(text):
    replacements = {
        "Ã³": "ó", "Ã©": "é", "Ã¡": "á", "Ã­": "í", "Ãº": "ú", "Ã±": "ñ",
        "Ã“": "Ó", "Ã‰": "É", "Ã": "Á", "Ã": "Í", "Ãš": "Ú", "Ã‘": "Ñ",
        "ï¬": "fi", "ï¬‚": "fl", "ﬁ": "fi", "ﬂ": "fl", "Â": "",
    }
    for bad, good in replacements.items():
        text = text.replace(bad, good)
    return text


def split_checklists(text):
    parts = re.split(r"(?=No\. Autogestion\s*\n?\s*\d+)", text)
    return [p for p in parts if re.search(r"No\. Autogestion", p, re.I)]


def date_from_checklist_text(text):
    m = re.search(r"Fecha de realizaci[oó]n\s+(\d{1,2})-([A-Za-zÁÉÍÓÚÑáéíóúñ]{3})-(\d{4})", fix_text(text), re.I)
    if not m:
        return None
    months = {
        "jan": 1, "ene": 1, "feb": 2, "mar": 3, "apr": 4, "abr": 4,
        "may": 5, "jun": 6, "jul": 7, "aug": 8, "ago": 8, "sep": 9,
        "oct": 10, "nov": 11, "dec": 12, "dic": 12,
    }
    month = months.get(m.group(2).lower()[:3])
    if not month:
        return None
    try:
        return date(int(m.group(3)), month, int(m.group(1)))
    except ValueError:
        return None


def sede_from_checklist_block(block):
    """Lee el campo 'Sede (Comfandi)' que trae cada checklist DENTRO del PDF.
    Esta es la fuente de verdad real: el nombre del archivo es solo una
    etiqueta puesta a mano y puede tener errores de tipeo, formatos de rango
    inconsistentes, o quedar mal copiado; el contenido del PDF no."""
    m = re.search(
        r"Sede \(Comfandi\)\s*(.*?)(?:Centro de costo|Fecha de realizaci[oó]n)",
        fix_text(block), re.S | re.I,
    )
    if not m:
        return ""
    return " ".join(x.strip() for x in fix_text(m.group(1)).splitlines() if x.strip())


def pdf_checklist_index():
    """Recorre cada PDF, separa sus checklists individuales y, para cada uno,
    lee del CONTENIDO (no del nombre del archivo) la sede y la fecha.

    Devuelve:
      - entries: lista plana de {archivo, sede_pdf, dia, fecha_iso} — una fila
        por checklist encontrado dentro de cualquier PDF. Esta es la base para
        calcular qué días quedaron cubiertos por sede, sin depender de si el
        nombre del archivo se pudo "limpiar" bien o no.
      - by_file: resumen por archivo (días, fechas, sedes vistas dentro del
        PDF, y el motivo si no se pudo leer nada) — se usa solo para mostrar
        evidencia/observaciones en el reporte, no para decidir cobertura.
    """
    entries = []
    by_file = {}
    if not PDF_TEXT_SUMMARY.exists():
        return entries, by_file
    for item in json.loads(PDF_TEXT_SUMMARY.read_text(encoding="utf-8")):
        name = item.get("Name", "")
        days = set()
        dates = []
        sedes_en_pdf = set()
        text_path = item.get("text_path") or ""
        source = "Sin fecha leída en PDF"
        if text_path and Path(text_path).exists():
            text = Path(text_path).read_text(encoding="utf-8", errors="ignore")
            for block in split_checklists(text):
                parsed = date_from_checklist_text(block)
                sede_pdf = sede_from_checklist_block(block)
                if sede_pdf:
                    sedes_en_pdf.add(sede_pdf)
                if parsed and parsed.year == YEAR and parsed.month == MONTH:
                    days.add(parsed.day)
                    dates.append(parsed.isoformat())
                    entries.append({
                        "archivo": name,
                        "sede_pdf": sede_pdf,
                        "dia": parsed.day,
                        "fecha_iso": parsed.isoformat(),
                    })
        if item.get("downloadedBytes") == 0 or int(item.get("Length") or 0) == 0:
            source = "PDF vacío / 0 bytes"
        elif item.get("text_error"):
            source = f"Error lectura PDF: {item.get('text_error')}"
        elif days:
            source = "Contenido PDF"
        by_file[name] = {
            "days": days,
            "dates": sorted(set(dates)),
            "sedes_en_pdf": sorted(sedes_en_pdf),
            "source": source,
        }
    return entries, by_file


def strip_day_list_parens(stem):
    """Elimina cualquier grupo entre paréntesis que sea una lista/rango de días,
    sin importar el separador: "(01A 11-14)", "(1-2-8-9-16-17-23-24-29-30)",
    "(27-28-29)", "(6-7-8)", "(11 )", etc. Antes solo se reconocía el patrón
    simple "num a num", por lo que estas variantes dejaban dígitos pegados al
    nombre de la sede y el archivo terminaba sin emparejar (score 0)."""

    def repl(match):
        content = match.group(1)
        cleaned = re.sub(r"(?i)\b(a|al|y)\b", " ", content)
        cleaned = re.sub(r"(?i)(\d)[a-z]+", r"\1", cleaned)
        cleaned = re.sub(r"[^\d]+", " ", cleaned).strip()
        tokens = cleaned.split()
        if tokens and all(t.isdigit() and 1 <= int(t) <= 31 for t in tokens):
            return " "
        return match.group(0)

    return re.sub(r"\(([^)]*)\)", repl, stem)


def strip_trailing_day_number(text):
    m = re.search(r"\b(\d{1,2})\s*$", text or "")
    if m and 1 <= int(m.group(1)) <= 31:
        return text[:m.start()].rstrip()
    return text


def site_from_filename(name):
    stem = re.sub(r"\.[^.]+$", "", name)
    stem = re.sub(r"^\s*\d+\s*[-.]?\s*", "", stem)
    stem = re.sub(r"(?i)\s*-\s*nombre\s+corregido\b.*$", "", stem)
    stem = re.sub(r"(?i)\bpor\s+firmas?\b.*$", "", stem)
    stem = re.sub(r"(?i)\bpor\s+firma\s+de\s+gestor\b.*$", "", stem)
    stem = re.sub(r"(?i)\bfirma\s+de\s+gestor\b.*$", "", stem)
    stem = re.sub(r"(?i)\bfirma(?:s)?\s+(?:e|y)\s+ind(?:ice)?\s+digital\b.*$", "", stem)
    stem = re.sub(r"(?i)\bpendiente\s+(?:por\s+)?firma\b.*$", "", stem)
    stem = re.sub(r"(?i)\bfalta\s+(?:el\s+)?[íi]ndice\b.*$", "", stem)
    stem = re.sub(r"(?i)\bsin\s+[íi]ndice\b.*$", "", stem)
    stem = re.sub(r"(?i)\bcon\s+[íi]ndice\b.*$", "", stem)
    stem = re.sub(r"(?i)\bpor\s+[íi]ndice\b.*$", "", stem)
    stem = re.sub(r"(?i)\bobservaci[oó]n(?:es)?\b.*$", "", stem)
    stem = re.sub(r"(?i)\bsoporte\b.*$", "", stem)
    stem = re.sub(r"(?i)\bcheck\s*list\b.*$", "", stem)
    stem = re.sub(r"(?i)\bchecklist\b.*$", "", stem)
    stem = strip_day_list_parens(stem)
    stem = re.sub(r"\(\s*\d{1,2}\s*(?:a|-)\s*\d{1,2}\s*\)", "", stem, flags=re.I)
    stem = re.sub(r"\(\s*\d{1,2}\s*\)", "", stem)
    stem = re.sub(r"\(([^)]*[A-Za-zÁÉÍÓÚÜÑáéíóúüñ][^)]*)\)", r" \1 ", stem)
    stem = re.sub(r"\b\d{1,2}\s*(?:a|-)\s*\d{1,2}\s*$", "", stem, flags=re.I)
    stem = strip_trailing_day_number(stem)
    stem = re.sub(r"\s+", " ", stem)
    stem = stem.strip(" -_.")
    return stem.strip()


def score_name(a, b):
    na, nb = normalize(a), normalize(b)
    if not na or not nb:
        return 0
    if na == nb:
        return 1
    seq = SequenceMatcher(None, na, nb).ratio()
    ta, tb = set(na.split()), set(nb.split())
    nums_a = {token for token in ta if token.isdigit()}
    nums_b = {token for token in tb if token.isdigit()}
    if (nums_a or nums_b) and nums_a != nums_b:
        return 0
    token = len(ta & tb) / max(len(ta | tb), 1)
    containment = len(ta & tb) / max(min(len(ta), len(tb)), 1)
    return max(seq, token, containment * 0.92)


def more_specific_site(candidate, current):
    if not current:
        return True
    return len(normalize(candidate).split()) > len(normalize(current).split())


def next_monday(day):
    return day + timedelta(days=(7 - day.weekday()) % 7)


def easter_sunday(year):
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)


def colombia_holidays(year):
    easter = easter_sunday(year)
    fixed = {
        date(year, 1, 1), date(year, 5, 1), date(year, 7, 20),
        date(year, 8, 7), date(year, 12, 8), date(year, 12, 25),
        easter - timedelta(days=3), easter - timedelta(days=2),
    }
    moved = [
        date(year, 1, 6), date(year, 3, 19), date(year, 6, 29),
        date(year, 8, 15), date(year, 10, 12), date(year, 11, 1),
        date(year, 11, 11), easter + timedelta(days=43),
        easter + timedelta(days=64), easter + timedelta(days=71),
    ]
    return fixed | {next_monday(day) for day in moved}


def day_number(value):
    if isinstance(value, (date,)):
        return value.day
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def expected_days_between(start_value, end_value):
    start = day_number(start_value)
    end = day_number(end_value)
    if start is None or end is None or start > end:
        return set(), set()
    holidays = colombia_holidays(YEAR)
    expected_days = set()
    excluded_days = set()
    for day in range(max(1, start), min(monthrange(YEAR, MONTH)[1], end) + 1):
        current = date(YEAR, MONTH, day)
        if current.weekday() == 6 or current in holidays:
            excluded_days.add(day)
        else:
            expected_days.add(day)
    return expected_days, excluded_days


def valid_workday(day):
    try:
        current = date(YEAR, MONTH, day)
    except ValueError:
        return False
    return current.weekday() != 6 and current not in colombia_holidays(YEAR)


def general_detail_days(schedule):
    return parse_days_text(schedule.get("dias_programados") or "")


def saturdays_not_programmed(start_value, end_value, expected_days):
    start = day_number(start_value)
    end = day_number(end_value)
    if start is None or end is None or start > end:
        return set()
    out = set()
    holidays = colombia_holidays(YEAR)
    for day in range(max(1, start), min(monthrange(YEAR, MONTH)[1], end) + 1):
        current = date(YEAR, MONTH, day)
        if current.weekday() == 5 and current not in holidays and day not in expected_days:
            out.add(day)
    return out


expected = json.loads(EXPECTED.read_text(encoding="utf-8"))
sp = json.loads(SP_FILES.read_text(encoding="utf-8"))
checklist_entries, pdf_days_by_name = pdf_checklist_index()
sites = expected["site_summary"]
site_lookup = {row["sede"]: row for row in sites}
site_lookup_normalized = {normalize(row["sede"]): row["sede"] for row in sites}
actual_by_site = {}
for row in sites:
    days = set()
    for d in re.findall(r"\d{1,2}(?:\s*-\s*\d{1,2})?", str(row.get("fecha_real_ejecucion") or "")):
        days.update(parse_days_text(d))
    if days:
        actual_by_site[row["sede"]] = days


def match_site(name_guess):
    """Empareja un nombre de sede (venga del archivo o del contenido del PDF)
    contra el listado real de sedes del cronograma."""
    if not name_guess:
        return None, 0
    exact = site_lookup_normalized.get(normalize(name_guess))
    if exact:
        return exact, 1
    best_site, best_score = None, 0
    for site in site_lookup:
        s = score_name(name_guess, site)
        if s > best_score or (abs(s - best_score) < 0.0001 and more_specific_site(site, best_site)):
            best_site, best_score = site, s
    return best_site, best_score


# --- Cobertura basada en el CONTENIDO del PDF (fuente de verdad) ---
# Cada checklist dentro de un PDF trae su propio campo "Sede (Comfandi)" y su
# propia "Fecha de realización". Un día queda cubierto para una sede si existe
# al menos un checklist, en cualquier PDF, cuya sede leída coincida con esa
# sede y cuya fecha caiga en ese día. Esto ya no depende de si el nombre del
# archivo se pudo interpretar bien (rangos "01 A 11", listas "27-28-29",
# "01A 11-14", errores de tipeo, etc.) — el nombre del archivo puede fallar y
# ya no afecta si el día queda marcado como cubierto.
covered_by_site = defaultdict(lambda: defaultdict(set))  # sede -> dia -> {archivos}
unmatched_checklists = []
for entry in checklist_entries:
    best_site, best_score = match_site(entry["sede_pdf"])
    if best_site and best_score >= 0.62:
        covered_by_site[best_site][entry["dia"]].add(entry["archivo"])
    else:
        unmatched_checklists.append({
            "archivo": entry["archivo"],
            "sede_leida_en_pdf": entry["sede_pdf"],
            "fecha": entry["fecha_iso"],
            "mejor_sede_candidata": best_site or "",
            "confianza": round(best_score, 3),
        })

# --- Info por archivo: solo para evidencia / control de calidad ---
# El nombre del archivo YA NO decide si un día cuenta como cubierto; solo se
# usa aquí para poder avisar si el nombre del archivo dice una sede distinta
# a la que realmente aparece adentro del PDF (posible archivo mal renombrado
# o subido en la carpeta equivocada).
file_rows = []
unmatched_files = []
for f in sp.get("files", []):
    name = f["Name"]
    file_site = site_from_filename(name)
    pdf_info = pdf_days_by_name.get(name, {"days": set(), "dates": [], "sedes_en_pdf": [], "source": "PDF no leído"})
    days = set(pdf_info["days"])
    best_site, best_score = match_site(file_site)
    sedes_en_pdf = pdf_info.get("sedes_en_pdf", [])
    sedes_en_pdf_matched = sorted({s for s, sc in (match_site(s) for s in sedes_en_pdf) if s and sc >= 0.62})
    nombre_vs_contenido_diferente = bool(
        best_site and sedes_en_pdf_matched and best_site not in sedes_en_pdf_matched
    )
    folder = f.get("ParentFolder", "").split("/")[-1]
    file_row = {
        "archivo": name,
        "carpeta": folder,
        "sede_archivo": file_site,
        "sede_match": best_site or "",
        "confianza_match": round(best_score, 3),
        "sede_leida_en_pdf": " | ".join(sedes_en_pdf),
        "nombre_vs_contenido_diferente": "Sí" if nombre_vs_contenido_diferente else "",
        "dias_archivo": ",".join(map(str, sorted(days))),
        "fechas_pdf": ",".join(pdf_info.get("dates", [])),
        "fuente_dias": pdf_info.get("source", "PDF no leído"),
        "creado_sharepoint": f.get("TimeCreated", ""),
        "modificado_sharepoint": f.get("TimeLastModified", ""),
        "tamano_bytes": int(f.get("Length") or 0),
        "url": sharepoint_file_view_url(f),
    }
    file_rows.append(file_row)
    if not (best_site and best_score >= 0.62 and days):
        unmatched_files.append(file_row)

detail_rows = []
summary_rows = []
for site in sorted(actual_by_site):
    days = actual_by_site.get(site, set())
    covered_map = covered_by_site.get(site, {})
    covered_days = {d for d in days if d in covered_map}
    exact_missing = sorted(days - covered_days)
    extra_file_days = sorted(set(covered_map.keys()) - days)
    if not exact_missing:
        site_status = "Completo"
        summary_missing = []
        observation = ""
    else:
        site_status = "Incompleto"
        summary_missing = exact_missing
        observation = ""
    matched_file_names = " | ".join(sorted({a for files in covered_map.values() for a in files}))
    for day in sorted(days):
        files = sorted(covered_map.get(day, set()))
        detail_status = "Cargado" if files else "Falta soporte"
        detail_rows.append({
            "sede": site,
            "ues": site_lookup[site].get("ues", ""),
            "ubicacion": site_lookup[site].get("ubicacion", ""),
            "fecha": f"{YEAR}-{MONTH:02d}-{day:02d}",
            "dia": day,
            "estado": detail_status,
            "archivo_soporte": " | ".join(files),
        })
    summary_rows.append({
        "sede": site,
        "ues": site_lookup[site].get("ues", ""),
        "ubicacion": site_lookup[site].get("ubicacion", ""),
        "dias_digitados": ",".join(map(str, sorted(days))),
        "cantidad_digitada": len(days),
        "dias_cubiertos_sharepoint": ",".join(map(str, sorted(covered_days))),
        "cantidad_cubierta": len(covered_days),
        "dias_faltantes": ",".join(map(str, summary_missing)),
        "cantidad_faltante": len(summary_missing),
        "dias_en_archivo_no_digitados": ",".join(map(str, extra_file_days)),
        "archivos_match": len({a for files in covered_map.values() for a in files}),
        "estado": site_status,
        "observacion": observation,
    })

calendar_rows = []
for site in sorted(site_lookup):
    schedule = site_lookup[site]
    start_value = schedule.get("cronog_fecha_inicio") or schedule.get("fecha_inicio")
    end_value = schedule.get("cronog_fecha_fin") or schedule.get("fecha_fin")
    source_range = "Cronog" if schedule.get("cronog_fecha_inicio") or schedule.get("cronog_fecha_fin") else "General"
    raw_detail_days = general_detail_days(schedule)
    detail_days = {day for day in raw_detail_days if valid_workday(day)}
    marked_not_counted = sorted(raw_detail_days - detail_days)
    range_days, excluded_days = expected_days_between(start_value, end_value)
    if detail_days:
        expected_days = detail_days
        source_expected = "General MAYO - programación diaria"
    else:
        expected_days = range_days
        source_expected = f"{source_range} - rango inicio/fin"
    saturday_not_programmed = saturdays_not_programmed(start_value, end_value, expected_days)
    digitized_days = actual_by_site.get(site, set())
    sharepoint_days = set(covered_by_site.get(site, {}).keys())
    missing_digitized = sorted(expected_days - digitized_days)
    digitized_without_support = sorted(digitized_days - sharepoint_days)
    expected_without_support = sorted(expected_days - sharepoint_days)
    support_not_digitized = sorted(sharepoint_days - digitized_days)
    digitized_outside_expected = sorted(digitized_days - expected_days)
    sharepoint_outside_expected = sorted(sharepoint_days - expected_days)
    alerts = []
    if missing_digitized:
        alerts.append(f"Cronograma: faltan {len(missing_digitized)} día(s) esperados por diligenciar")
    if digitized_without_support:
        alerts.append(f"SharePoint: faltan {len(digitized_without_support)} soporte(s) de días ya digitados")
    if support_not_digitized:
        alerts.append(f"Cruce: hay {len(support_not_digitized)} día(s) cargados en SharePoint sin estar digitados")
    if digitized_outside_expected:
        alerts.append(f"Cronograma: hay {len(digitized_outside_expected)} fecha(s) digitadas fuera del esperado")
    if sharepoint_outside_expected:
        alerts.append(f"SharePoint: hay {len(sharepoint_outside_expected)} soporte(s) cargados fuera del rango esperado")
    if not expected_days:
        alerts.append("Revisar fecha inicio y fecha fin")
    calendar_rows.append({
        "sede": site,
        "ues": schedule.get("ues", ""),
        "ubicacion": schedule.get("ubicacion", ""),
        "fecha_inicio": start_value,
        "fecha_fin": end_value,
        "fuente_inicio_fin": source_range,
        "fuente_esperado": source_expected,
        "general_fecha_inicio": schedule.get("fecha_inicio", ""),
        "general_fecha_fin": schedule.get("fecha_fin", ""),
        "cuadritos_general": ",".join(map(str, sorted(raw_detail_days))),
        "marcados_no_contados": ",".join(map(str, marked_not_counted)),
        "dias_programados_general": ",".join(map(str, sorted(detail_days))),
        "sabados_sin_visita": ",".join(map(str, sorted(saturday_not_programmed))),
        "cantidad_sabados_sin_visita": len(saturday_not_programmed),
        "dias_excluidos_domingo_festivo": ",".join(map(str, sorted(excluded_days))),
        "dias_esperados": ",".join(map(str, sorted(expected_days))),
        "cantidad_esperada": len(expected_days),
        "dias_digitados": ",".join(map(str, sorted(digitized_days))),
        "cantidad_digitada": len(digitized_days),
        "dias_sharepoint": ",".join(map(str, sorted(sharepoint_days))),
        "cantidad_sharepoint": len(sharepoint_days),
        "dias_falta_diligenciar": ",".join(map(str, missing_digitized)),
        "cantidad_falta_diligenciar": len(missing_digitized),
        "dias_digitados_falta_cargar": ",".join(map(str, digitized_without_support)),
        "cantidad_digitada_falta_cargar": len(digitized_without_support),
        "dias_esperados_sin_sharepoint": ",".join(map(str, expected_without_support)),
        "cantidad_esperada_sin_sharepoint": len(expected_without_support),
        "dias_sharepoint_no_digitados": ",".join(map(str, support_not_digitized)),
        "cantidad_sharepoint_no_digitada": len(support_not_digitized),
        "dias_digitados_fuera_esperado": ",".join(map(str, digitized_outside_expected)),
        "dias_sharepoint_fuera_esperado": ",".join(map(str, sharepoint_outside_expected)),
        "alerta": " | ".join(alerts) if alerts else "Sin alertas",
        "estado": "Completo" if not alerts else "Incompleto",
    })


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, ws.max_column + 1):
        values = [ws.cell(row, col).value for row in range(1, min(ws.max_row, 120) + 1)]
        width = min(max(len(str(v)) if v is not None else 0 for v in values) + 2, 60)
        ws.column_dimensions[get_column_letter(col)].width = max(width, 10)


def write_rows(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    style_sheet(ws)


OUT.parent.mkdir(parents=True, exist_ok=True)
wb = Workbook()
ws = wb.active
ws.title = "Resumen"
total_digitados = sum(r["cantidad_digitada"] for r in summary_rows)
total_cubiertos = sum(r["cantidad_cubierta"] for r in summary_rows)
total_faltantes = sum(r["cantidad_faltante"] for r in summary_rows)
overview = [
    {"concepto": "Archivos encontrados en SharePoint", "resultado": len(file_rows)},
    {"concepto": "Sedes con fechas digitadas en Cronog Mayo 2026", "resultado": len(summary_rows)},
    {"concepto": "Días/checklists digitados", "resultado": total_digitados},
    {"concepto": "Días/checklists cubiertos por archivos cargados", "resultado": total_cubiertos},
    {"concepto": "Días/checklists faltantes", "resultado": total_faltantes},
    {"concepto": "Sedes completas", "resultado": sum(1 for r in summary_rows if r["estado"] == "Completo")},
    {"concepto": "Sedes incompletas", "resultado": sum(1 for r in summary_rows if r["estado"] != "Completo")},
    {"concepto": "Archivos sin match confiable por nombre (solo referencia)", "resultado": len(unmatched_files)},
    {"concepto": "Checklists con sede no reconocida dentro del PDF", "resultado": len(unmatched_checklists)},
    {"concepto": "Nota", "resultado": "La cobertura se calcula leyendo 'Sede (Comfandi)' y 'Fecha de realización' DENTRO de cada checklist del PDF. El nombre del archivo ya no decide si un día está cubierto; solo se usa para detectar archivos mal renombrados (ver hoja 'Nombre vs contenido')."},
]
write_rows(ws, ["concepto", "resultado"], overview)

ws = wb.create_sheet("Resumen por sede")
write_rows(ws, [
    "sede", "ues", "ubicacion", "dias_digitados", "cantidad_digitada",
    "dias_cubiertos_sharepoint", "cantidad_cubierta", "dias_faltantes",
    "cantidad_faltante", "dias_en_archivo_no_digitados", "archivos_match", "estado", "observacion"
], summary_rows)

ws = wb.create_sheet("Detalle por fecha")
write_rows(ws, ["sede", "ues", "ubicacion", "fecha", "dia", "estado", "archivo_soporte"], detail_rows)

ws = wb.create_sheet("Archivos SharePoint")
write_rows(ws, ["archivo", "carpeta", "sede_archivo", "sede_match", "confianza_match", "sede_leida_en_pdf", "nombre_vs_contenido_diferente", "dias_archivo", "fechas_pdf", "fuente_dias", "creado_sharepoint", "modificado_sharepoint", "tamano_bytes", "url"], file_rows)

ws = wb.create_sheet("Archivos sin match")
write_rows(ws, ["archivo", "carpeta", "sede_archivo", "sede_match", "confianza_match", "sede_leida_en_pdf", "nombre_vs_contenido_diferente", "dias_archivo", "fechas_pdf", "fuente_dias", "creado_sharepoint", "modificado_sharepoint", "tamano_bytes", "url"], unmatched_files)

ws = wb.create_sheet("Checklists sin sede reconocida")
write_rows(ws, ["archivo", "sede_leida_en_pdf", "fecha", "mejor_sede_candidata", "confianza"], unmatched_checklists)

ws = wb.create_sheet("Validación calendario")
write_rows(ws, [
    "sede", "ues", "ubicacion", "fecha_inicio", "fecha_fin", "fuente_inicio_fin",
    "fuente_esperado",
    "general_fecha_inicio", "general_fecha_fin",
    "cuadritos_general", "marcados_no_contados",
    "dias_programados_general", "sabados_sin_visita", "cantidad_sabados_sin_visita",
    "dias_excluidos_domingo_festivo", "dias_esperados", "cantidad_esperada",
    "dias_digitados", "cantidad_digitada", "dias_sharepoint", "cantidad_sharepoint",
    "dias_falta_diligenciar", "cantidad_falta_diligenciar",
    "dias_digitados_falta_cargar", "cantidad_digitada_falta_cargar",
    "dias_esperados_sin_sharepoint", "cantidad_esperada_sin_sharepoint",
    "dias_sharepoint_no_digitados", "cantidad_sharepoint_no_digitada",
    "dias_digitados_fuera_esperado", "dias_sharepoint_fuera_esperado", "alerta", "estado",
], calendar_rows)

wb.save(OUT)
print(json.dumps({
    "sharepoint_files": len(file_rows),
    "sedes_digitadas": len(summary_rows),
    "dias_digitados": total_digitados,
    "dias_cubiertos": total_cubiertos,
    "dias_faltantes": total_faltantes,
    "sedes_completas": sum(1 for r in summary_rows if r["estado"] == "Completo"),
    "sedes_incompletas": sum(1 for r in summary_rows if r["estado"] != "Completo"),
    "archivos_sin_match": len(unmatched_files),
    "checklists_sede_no_reconocida": len(unmatched_checklists),
    "sedes_alerta_calendario": sum(1 for row in calendar_rows if row["estado"] == "Incompleto"),
    "checklists_esperados_calendario": sum(row["cantidad_esperada"] for row in calendar_rows),
    "checklists_falta_diligenciar": sum(row["cantidad_falta_diligenciar"] for row in calendar_rows),
    "checklists_digitados_falta_cargar": sum(row["cantidad_digitada_falta_cargar"] for row in calendar_rows),
    "output": str(OUT),
}, ensure_ascii=False, indent=2))
