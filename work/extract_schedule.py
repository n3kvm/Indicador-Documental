import json
import os
import re
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


INPUT = Path(os.environ.get("CRONOGRAMA_PATH", r"D:\Datos\OneDrive - BRILLASEO SAS\Descargas\05-Cronograma General Mtto Prog Mayo 2026 (Planeación) (1).xlsx"))
OUT = Path(os.environ.get("EXPECTED_SCHEDULE_PATH", "work/expected_schedule.json"))
REPORT = Path(os.environ.get("SCHEDULE_AUDIT_OUTPUT_PATH", r"C:\Users\DELL\Documents\Codex\2026-06-03\files-mentioned-by-the-user-05\outputs\auditoria_soportes_mayo_2026.xlsx"))
YEAR = int(os.environ.get("SHAREPOINT_ANIO", "2026"))
MONTH = int(os.environ.get("SHAREPOINT_MES", "5"))
MONTH_NAME = os.environ.get("SHAREPOINT_MES_NOMBRE", "").upper()


def clean(value):
    if value is None:
        return ""
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def parse_days(value):
    if value is None:
        return []
    if isinstance(value, (int, float)):
        return [int(value)]
    return [int(x) for x in re.findall(r"\d+", str(value))]


def norm(value):
    return clean(value).upper()


def find_row_with_headers(sheet, required, max_rows=12):
    required = [r.upper() for r in required]
    for row in range(1, min(sheet.max_row, max_rows) + 1):
        values = [norm(sheet.cell(row, col).value) for col in range(1, sheet.max_column + 1)]
        joined = " | ".join(values)
        if all(req in joined for req in required):
            return row
    return None


def find_general_sheet(workbook):
    candidates = []
    for sheet in workbook.worksheets:
        header_row = find_row_with_headers(sheet, ["SEDE", "FECHA INICIO", "FECHA FIN"])
        if not header_row:
            continue
        day_row = header_row - 1
        day_count = 0
        if day_row >= 1:
            for col in range(1, sheet.max_column + 1):
                value = sheet.cell(day_row, col).value
                if isinstance(value, (int, float)) and 1 <= int(value) <= 31:
                    day_count += 1
        score = day_count
        if "GENERAL" in norm(sheet.title):
            score += 20
        if MONTH_NAME and MONTH_NAME in norm(sheet.title):
            score += 100
        candidates.append((score, sheet, header_row, day_row))
    if not candidates:
        raise KeyError(f"No encontré hoja general de cronograma. Hojas disponibles: {workbook.sheetnames}")
    if MONTH_NAME and not any(MONTH_NAME in norm(item[1].title) for item in candidates):
        raise KeyError(f"No encontré hoja General para {MONTH_NAME}. Hojas disponibles: {workbook.sheetnames}")
    return sorted(candidates, key=lambda item: item[0], reverse=True)[0]


def find_cronog_sheet(workbook):
    candidates = []
    for sheet in workbook.worksheets:
        header_row = find_row_with_headers(sheet, ["SEDES PROGRAMADAS", "FECHA REAL DE EJECUCIÓN", "TOTAL AUTOGESTIONES"])
        if not header_row:
            continue
        score = 10
        if "CRONOG" in norm(sheet.title):
            score += 20
        if MONTH_NAME and MONTH_NAME in norm(sheet.title):
            score += 100
        candidates.append((score, sheet, header_row))
    if not candidates:
        raise KeyError(f"No encontré hoja Cronog con fechas reales. Hojas disponibles: {workbook.sheetnames}")
    if MONTH_NAME and not any(MONTH_NAME in norm(item[1].title) for item in candidates):
        raise KeyError(f"No encontré hoja Cronog para {MONTH_NAME}. Hojas disponibles: {workbook.sheetnames}")
    return sorted(candidates, key=lambda item: item[0], reverse=True)[0]


def header_map(sheet, header_row):
    out = {}
    for col in range(1, sheet.max_column + 1):
        key = norm(sheet.cell(header_row, col).value)
        if key:
            out[key] = col
    return out


def find_col(headers, *names):
    for name in names:
        wanted = name.upper()
        if wanted in headers:
            return headers[wanted]
    for key, col in headers.items():
        if any(name.upper() in key for name in names):
            return col
    return None


wb = openpyxl.load_workbook(INPUT, data_only=True)
_, ws, general_header_row, day_row = find_general_sheet(wb)
_, cron, cron_header_row = find_cronog_sheet(wb)
general_headers = header_map(ws, general_header_row)
cron_headers = header_map(cron, cron_header_row)
print(json.dumps({
    "general_sheet": ws.title,
    "general_header_row": general_header_row,
    "cronog_sheet": cron.title,
    "cronog_header_row": cron_header_row,
}, ensure_ascii=False))

date_cols = []
for col in range(1, ws.max_column + 1):
    day = ws.cell(day_row, col).value
    if isinstance(day, (int, float)) and 1 <= int(day) <= 31:
        date_cols.append((col, int(day), clean(ws.cell(general_header_row, col).value)))

col_sede = find_col(general_headers, "SEDE", "SEDE  ")
col_ues = find_col(general_headers, "UES")
col_ubicacion = find_col(general_headers, "UBICACIÓN", "UBICACION")
col_frecuencia = find_col(general_headers, "FRECUENCIA")
col_responsable = find_col(general_headers, "RESPONSABLE")
col_tipo = find_col(general_headers, "TIPO")
col_horas_prog = find_col(general_headers, "HORAS PROGRAMADAS MES", "Horas Programadas Mes")
col_inicio = find_col(general_headers, "FECHA INICIO")
col_fin = find_col(general_headers, "FECHA FIN")
col_dias = find_col(general_headers, "DIAS", "DÍAS")
col_grupo = find_col(general_headers, "GRUPO")

events = []
actual_events = []
site_summary = []
for row in range(general_header_row + 1, ws.max_row + 1):
    sede = clean(ws.cell(row, col_sede).value)
    if not sede:
        continue
    base = {
        "sede": sede,
        "ues": clean(ws.cell(row, col_ues).value) if col_ues else "",
        "ubicacion": clean(ws.cell(row, col_ubicacion).value) if col_ubicacion else "",
        "frecuencia": clean(ws.cell(row, col_frecuencia).value) if col_frecuencia else "",
        "responsable": clean(ws.cell(row, col_responsable).value) if col_responsable else "",
        "tipo": clean(ws.cell(row, col_tipo).value) if col_tipo else "",
        "horas_programadas_mes": ws.cell(row, col_horas_prog).value if col_horas_prog else "",
        "fecha_inicio": ws.cell(row, col_inicio).value if col_inicio else "",
        "fecha_fin": ws.cell(row, col_fin).value if col_fin else "",
        "dias_cronograma": ws.cell(row, col_dias).value if col_dias else "",
        "grupo": clean(ws.cell(row, col_grupo).value) if col_grupo else "",
    }
    programmed = []
    excluded = []
    for col, day, dow in date_cols:
        code = clean(ws.cell(row, col).value)
        if code and code.upper() != "F":
            record = dict(base)
            record.update({"dia": day, "fecha": f"{YEAR}-{MONTH:02d}-{day:02d}", "dia_semana": dow, "codigo_programacion": code})
            events.append(record)
            programmed.append(day)
        elif code.upper() == "F":
            excluded.append(day)
    site_summary.append({
        **base,
        "checklists_esperados": len(programmed),
        "dias_programados": ",".join(map(str, programmed)),
        "dias_excluidos_f": ",".join(map(str, excluded)),
    })

cron_summary = {}
cron_col_sede = find_col(cron_headers, "SEDES PROGRAMADAS")
cron_col_duracion = find_col(cron_headers, "DURACION DIAS", "DURACIÓN DIAS")
cron_col_inicio = find_col(cron_headers, "FECHA INICIO")
cron_col_fin = find_col(cron_headers, "FECHA FIN")
cron_col_fecha_real = find_col(cron_headers, "FECHA REAL DE EJECUCIÓN", "FECHA REAL DE EJECUCION")
cron_col_horas = find_col(cron_headers, "HORAS REALES EJ")
cron_col_autog = find_col(cron_headers, "TOTAL AUTOGESTIONES")

for row in range(cron_header_row + 1, cron.max_row + 1):
    sede = clean(cron.cell(row, cron_col_sede).value)
    if sede and not sede.upper().startswith("ZONA"):
        cron_summary[sede] = {
            "duracion_dias": cron.cell(row, cron_col_duracion).value if cron_col_duracion else "",
            "cronog_fecha_inicio": cron.cell(row, cron_col_inicio).value if cron_col_inicio else "",
            "cronog_fecha_fin": cron.cell(row, cron_col_fin).value if cron_col_fin else "",
            "fecha_real_ejecucion": clean(cron.cell(row, cron_col_fecha_real).value) if cron_col_fecha_real else "",
            "dias_ejecutados_reportados": len(parse_days(cron.cell(row, cron_col_fecha_real).value)) if cron_col_fecha_real else 0,
            "horas_reales": cron.cell(row, cron_col_horas).value if cron_col_horas else "",
            "total_autogestiones": cron.cell(row, cron_col_autog).value if cron_col_autog else "",
        }

for row in site_summary:
    info = cron_summary.get(row["sede"], {})
    row.update(info)
    row["diferencia_esperado_vs_duracion"] = row["checklists_esperados"] - int(row["duracion_dias"] or 0) if info.get("duracion_dias") is not None else None
    row["diferencia_esperado_vs_ejecutado_reportado"] = row["checklists_esperados"] - int(row["dias_ejecutados_reportados"] or 0) if info else None
    for day in parse_days(row.get("fecha_real_ejecucion")):
        actual_events.append({
            "sede": row["sede"],
            "ues": row["ues"],
            "ubicacion": row["ubicacion"],
            "frecuencia": row["frecuencia"],
            "responsable": row["responsable"],
            "tipo": row["tipo"],
            "fecha": f"{YEAR}-{MONTH:02d}-{day:02d}",
            "dia": day,
            "fuente": "Cronog Mayo 2026 - Fecha real de ejecución",
        })

errors = []
for sheet in wb.worksheets:
    for cells in sheet.iter_rows():
        for cell in cells:
            value = cell.value
            if isinstance(value, str) and value.startswith("#"):
                errors.append({"hoja": sheet.title, "celda": cell.coordinate, "valor": value})

OUT.parent.mkdir(parents=True, exist_ok=True)
OUT.write_text(json.dumps({"events": events, "actual_events": actual_events, "site_summary": site_summary, "errors": errors}, ensure_ascii=False, indent=2), encoding="utf-8")


def write_rows(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    style_sheet(ws)


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, ws.max_column + 1):
        values = [ws.cell(row, col).value for row in range(1, min(ws.max_row, 80) + 1)]
        width = min(max(len(str(v)) if v is not None else 0 for v in values) + 2, 48)
        ws.column_dimensions[get_column_letter(col)].width = max(width, 10)


REPORT.parent.mkdir(parents=True, exist_ok=True)
out_wb = Workbook()
summary_ws = out_wb.active
summary_ws.title = "Resumen"
summary_rows = [
    ("Archivo revisado", str(INPUT)),
    ("Estado SharePoint", "Pendiente: no fue posible leer el enlace desde esta sesión."),
    ("Sedes en cronograma", len(site_summary)),
    ("Checklists esperados según General MAYO", len(events)),
    ("Checklists según fecha real de ejecución", len(actual_events)),
    ("Sedes con fecha real vacía o menor a lo programado", sum(1 for r in site_summary if (r.get("diferencia_esperado_vs_ejecutado_reportado") or 0) > 0)),
    ("Celdas con error visible en libro", len(errors)),
]
summary_ws.append(["Concepto", "Resultado"])
for item in summary_rows:
    summary_ws.append(item)
summary_ws.append([])
summary_ws.append(["Cómo usar", "Exporta el listado de archivos de SharePoint o descarga la carpeta SOPORTE DE VISITAS; con esa lista se compara contra las hojas de esperado por fecha."])
style_sheet(summary_ws)

ws = out_wb.create_sheet("Esperado plan")
write_rows(ws, ["sede", "ues", "ubicacion", "frecuencia", "responsable", "tipo", "fecha", "dia", "dia_semana", "codigo_programacion"], events)

ws = out_wb.create_sheet("Esperado ejecucion")
write_rows(ws, ["sede", "ues", "ubicacion", "frecuencia", "responsable", "tipo", "fecha", "dia", "fuente"], actual_events)

ws = out_wb.create_sheet("Resumen sedes")
write_rows(ws, [
    "sede", "ues", "ubicacion", "frecuencia", "responsable", "tipo", "checklists_esperados",
    "dias_programados", "fecha_inicio", "fecha_fin", "cronog_fecha_inicio", "cronog_fecha_fin",
    "duracion_dias", "fecha_real_ejecucion", "dias_ejecutados_reportados",
    "diferencia_esperado_vs_duracion", "diferencia_esperado_vs_ejecutado_reportado",
], site_summary)

alert_rows = [
    r for r in site_summary
    if r.get("duracion_dias") is None
    or r.get("diferencia_esperado_vs_ejecutado_reportado") not in (0, None)
    or r.get("diferencia_esperado_vs_duracion") not in (0, None)
]
ws = out_wb.create_sheet("Alertas cronograma")
write_rows(ws, [
    "sede", "ues", "ubicacion", "checklists_esperados", "dias_programados",
    "duracion_dias", "fecha_real_ejecucion", "dias_ejecutados_reportados",
    "diferencia_esperado_vs_duracion", "diferencia_esperado_vs_ejecutado_reportado",
], alert_rows)

ws = out_wb.create_sheet("Errores libro")
write_rows(ws, ["hoja", "celda", "valor"], errors)

ws = out_wb.create_sheet("Carga SharePoint")
write_rows(ws, [
    "fecha", "sede", "archivo_sharepoint", "encontrado", "corresponde_a_sede",
    "sin_informacion_pendiente", "observaciones",
], [])

out_wb.save(REPORT)
print(json.dumps({
    "sedes": len(site_summary),
    "checklists_esperados": len(events),
    "checklists_fecha_real": len(actual_events),
    "sin_match_cronog": sum(1 for r in site_summary if r.get("duracion_dias") is None),
    "diferencias_duracion": sum(1 for r in site_summary if r.get("diferencia_esperado_vs_duracion") not in (0, None)),
    "diferencias_ejecutado": sum(1 for r in site_summary if r.get("diferencia_esperado_vs_ejecutado_reportado") not in (0, None)),
    "errores_libro": len(errors),
    "reporte": str(REPORT),
}, ensure_ascii=False, indent=2))
