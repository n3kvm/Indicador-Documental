import csv
import json
import os
import re
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl


COMPARE_OUTPUT = Path(os.environ.get("COMPARE_OUTPUT_PATH", "outputs/validacion_soportes_sharepoint.xlsx"))
ZOHO_CSV = Path(os.environ.get("ZOHO_OUTPUT_CSV_PATH", "work/zoho_autogestiones.csv"))
CSV_PATH = Path(os.environ.get("APP_AUTOGESTIONES_CSV_PATH") or (str(ZOHO_CSV) if ZOHO_CSV.exists() else r"D:\Datos\OneDrive - BRILLASEO SAS\Descargas\Master Autogestiones.csv"))
OUTPUT = Path(os.environ.get("APP_UPLOAD_VALIDATION_PATH", "work/app_upload_validation.json"))
YEAR = int(os.environ.get("SHAREPOINT_ANIO", "2026"))
MONTH = int(str(os.environ.get("SHAREPOINT_CARPETA_MES", os.environ.get("SHAREPOINT_MES", "06"))).split("-", 1)[0])
SOURCE_LABEL = "Zoho Creator API" if CSV_PATH.resolve() == ZOHO_CSV.resolve() else "CSV"


MONTHS = {
    "jan": 1, "ene": 1, "january": 1, "enero": 1,
    "feb": 2, "february": 2, "febrero": 2,
    "mar": 3, "march": 3, "marzo": 3,
    "apr": 4, "abr": 4, "april": 4, "abril": 4,
    "may": 5, "mayo": 5,
    "jun": 6, "june": 6, "junio": 6,
    "jul": 7, "july": 7, "julio": 7,
    "aug": 8, "ago": 8, "august": 8, "agosto": 8,
    "sep": 9, "sept": 9, "september": 9, "septiembre": 9,
    "oct": 10, "october": 10, "octubre": 10,
    "nov": 11, "november": 11, "noviembre": 11,
    "dec": 12, "dic": 12, "december": 12, "diciembre": 12,
}


def normalize_text(value):
    text = str(value or "").strip()
    text = text.replace("Ã³", "o").replace("Ã©", "e").replace("Ã­", "i").replace("Ã¡", "a").replace("Ãº", "u")
    text = text.replace("Ã‘", "N").replace("Ã±", "n").replace("Ã“", "O")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def site_key(value):
    text = normalize_text(value)
    stopwords = {
        "sede", "de", "del", "la", "el", "y", "en",
    }
    tokens = [token for token in text.split() if token not in stopwords]
    return " ".join(tokens) or text


def split_days(value):
    days = []
    for part in re.findall(r"\d+", str(value or "")):
        day = int(part)
        if 1 <= day <= 31 and day not in days:
            days.append(day)
    return sorted(days)


def parse_date(value):
    text = str(value or "").strip()
    if not text:
        return None
    text = text.split()[0]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    match = re.match(r"^(\d{1,2})[-/ ]([A-Za-zÁÉÍÓÚáéíóúñÑ]+)[-/ ](\d{4})$", text)
    if match:
        day = int(match.group(1))
        month_name = normalize_text(match.group(2)).replace(" ", "")
        year = int(match.group(3))
        month = MONTHS.get(month_name[:3]) or MONTHS.get(month_name)
        if month:
            return datetime(year, month, day).date()
    return None


def read_csv_rows(path):
    if not path.exists():
        return []
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with path.open("r", encoding=encoding, newline="") as fh:
                return list(csv.DictReader(fh))
        except UnicodeDecodeError:
            continue
    with path.open("r", encoding="utf-8", errors="replace", newline="") as fh:
        return list(csv.DictReader(fh))


def find_col(headers, *needles):
    normalized = {header: normalize_text(header) for header in headers}
    for header, value in normalized.items():
        if all(needle in value for needle in needles):
            return header
    return ""


def rows_from_sheet(workbook, sheet_name):
    ws = workbook[sheet_name]
    headers = [str(cell.value or "") for cell in ws[1]]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and str(value).strip() for value in values):
            continue
        rows.append({header: ("" if value is None else value) for header, value in zip(headers, values)})
    return rows


def best_site_match(csv_site, expected_rows):
    key = site_key(csv_site)
    exact = [row for row in expected_rows if site_key(row.get("sede")) == key]
    if exact:
        return exact[0]
    csv_tokens = set(key.split())
    best = None
    best_score = 0.0
    for row in expected_rows:
        expected_key = site_key(row.get("sede"))
        expected_tokens = set(expected_key.split())
        if not csv_tokens or not expected_tokens:
            continue
        score = len(csv_tokens & expected_tokens) / max(len(csv_tokens), len(expected_tokens))
        if score > best_score:
            best_score = score
            best = row
    return best if best_score >= 0.72 else None


def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    if not COMPARE_OUTPUT.exists():
        OUTPUT.write_text(json.dumps({
            "enabled": False,
            "message": f"No existe el archivo base de validacion: {COMPARE_OUTPUT}",
            "source": str(CSV_PATH),
            "summary": {},
            "rows": [],
            "detail": [],
            "unmatched": [],
        }, ensure_ascii=False, indent=2), encoding="utf-8")
        return

    wb = openpyxl.load_workbook(COMPARE_OUTPUT, data_only=True)
    sheet_name = "Validación calendario" if "Validación calendario" in wb.sheetnames else "ValidaciÃ³n calendario"
    expected_rows = rows_from_sheet(wb, sheet_name)

    raw_rows = read_csv_rows(CSV_PATH)
    if not raw_rows:
        OUTPUT.write_text(json.dumps({
            "enabled": False,
            "message": f"No se encontraron autogestiones desde {SOURCE_LABEL}: {CSV_PATH}",
            "source": str(CSV_PATH),
            "source_type": SOURCE_LABEL,
            "summary": {
                "sedes_revisadas": len(expected_rows),
                "checklists_esperados": sum(len(split_days(r.get("dias_programados_general") or r.get("dias_esperados"))) for r in expected_rows),
                "cargados_aplicativo": 0,
                "faltan_aplicativo": 0,
                "cobertura_aplicativo": 0,
            },
            "rows": [],
            "detail": [],
            "unmatched": [],
        }, ensure_ascii=False, indent=2), encoding="utf-8")
        return

    headers = list(raw_rows[0].keys())
    col_auto = find_col(headers, "autogestion")
    col_site = find_col(headers, "nombre", "sede") or find_col(headers, "sede")
    col_ues = find_col(headers, "ues")
    col_done = find_col(headers, "fecha", "realizaci")
    col_registered = find_col(headers, "fecha", "registro")

    uploads = defaultdict(lambda: {"days": set(), "items": [], "ues": ""})
    unmatched = []
    for row in raw_rows:
        date = parse_date(row.get(col_done, ""))
        if not date or date.year != YEAR or date.month != MONTH:
            continue
        match = best_site_match(row.get(col_site, ""), expected_rows)
        item = {
            "autogestion": row.get(col_auto, ""),
            "sede_csv": row.get(col_site, ""),
            "ues_csv": row.get(col_ues, ""),
            "fecha_realizacion": date.isoformat(),
            "dia": date.day,
            "fecha_registro": row.get(col_registered, ""),
        }
        if not match:
            unmatched.append(item)
            continue
        key = str(match.get("sede", ""))
        uploads[key]["days"].add(date.day)
        uploads[key]["ues"] = row.get(col_ues, "")
        uploads[key]["items"].append(item)

    output_rows = []
    detail = []
    total_expected = 0
    total_uploaded = 0
    total_missing = 0

    for row in expected_rows:
        expected_days = split_days(row.get("dias_programados_general") or row.get("dias_esperados"))
        uploaded_days = sorted(uploads[str(row.get("sede", ""))]["days"])
        missing_days = [day for day in expected_days if day not in uploaded_days]
        outside_days = [day for day in uploaded_days if day not in expected_days]
        autos = sorted({str(item.get("autogestion", "")).strip() for item in uploads[str(row.get("sede", ""))]["items"] if str(item.get("autogestion", "")).strip()})
        status = "Completo" if not missing_days else "Incompleto"
        if status == "Completo" and outside_days:
            status = "Completo con observacion"
        alert = "Faltan dias por subir al aplicativo" if missing_days else ("Tiene dias cargados fuera de la programacion" if outside_days else "Sin faltantes")

        total_expected += len(expected_days)
        total_uploaded += len([day for day in uploaded_days if day in expected_days])
        total_missing += len(missing_days)

        output_rows.append({
            "sede": row.get("sede", ""),
            "ues": row.get("ues", "") or uploads[str(row.get("sede", ""))]["ues"],
            "ubicacion": row.get("ubicacion", ""),
            "fecha_inicio": row.get("fecha_inicio", ""),
            "fecha_fin": row.get("fecha_fin", ""),
            "fuente_esperado": row.get("fuente_esperado", ""),
            "dias_esperados_general": ",".join(map(str, expected_days)),
            "cantidad_esperada_general": len(expected_days),
            "dias_subidos_aplicativo": ",".join(map(str, uploaded_days)),
            "cantidad_subida_aplicativo": len(uploaded_days),
            "dias_falta_subir_aplicativo": ",".join(map(str, missing_days)),
            "cantidad_falta_subir_aplicativo": len(missing_days),
            "dias_fuera_programacion_aplicativo": ",".join(map(str, outside_days)),
            "cantidad_fuera_programacion_aplicativo": len(outside_days),
            "autogestiones_csv": ", ".join(autos),
            "cantidad_autogestiones_csv": len(autos),
            "estado": status,
            "alerta": alert,
        })

        for day in expected_days:
            detail.append({
                "sede": row.get("sede", ""),
                "ues": row.get("ues", ""),
                "dia": day,
                "esperado_general": "Si",
                "subido_aplicativo": "Si" if day in uploaded_days else "No",
                "estado": "Subido" if day in uploaded_days else "Falta subir",
            })

    summary = {
        "sedes_revisadas": len(output_rows),
        "checklists_esperados": total_expected,
        "cargados_aplicativo": total_uploaded,
        "faltan_aplicativo": total_missing,
        "cobertura_aplicativo": round((total_uploaded / total_expected * 100) if total_expected else 0, 1),
        "autogestiones_csv_periodo": sum(len(v["items"]) for v in uploads.values()) + len(unmatched),
        "sin_match": len(unmatched),
    }

    OUTPUT.write_text(json.dumps({
        "enabled": True,
        "message": f"Validacion de aplicativo generada desde {SOURCE_LABEL}.",
        "source": str(CSV_PATH),
        "source_type": SOURCE_LABEL,
        "summary": summary,
        "rows": output_rows,
        "detail": detail,
        "unmatched": unmatched,
    }, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
