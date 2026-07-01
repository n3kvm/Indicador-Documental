import json
import os
from pathlib import Path

import openpyxl


BASE = Path(os.environ.get("COMPARE_OUTPUT_PATH", r"C:\Users\DELL\Documents\Codex\2026-06-03\files-mentioned-by-the-user-05\outputs\validacion_soportes_sharepoint_mayo_2026.xlsx"))
PILLARS = Path(os.environ.get("PILLARS_VALIDATION_PATH", "work/pillars_hours_validation.json"))
APP_UPLOADS = Path(os.environ.get("APP_UPLOAD_VALIDATION_PATH", "work/app_upload_validation.json"))
OUTPUT = Path(os.environ.get("DASHBOARD_OUTPUT_PATH", r"C:\Users\DELL\Documents\Codex\2026-06-03\files-mentioned-by-the-user-05\outputs\dashboard_soportes_pilares_horas_mayo_2026.html"))
ANIO = os.environ.get("SHAREPOINT_ANIO", "2026")
CARPETA_MES = os.environ.get("SHAREPOINT_CARPETA_MES", "05-MAYO")
REFRESH_ENDPOINT = os.environ.get("DASHBOARD_REFRESH_ENDPOINT", "")
GITHUB_REPOSITORY = os.environ.get("GITHUB_REPOSITORY", "")
GITHUB_WORKFLOW_FILE = os.environ.get("DASHBOARD_GITHUB_WORKFLOW_FILE", "dashboard-github-pages.yml")
GITHUB_BRANCH = os.environ.get("DASHBOARD_GITHUB_BRANCH", os.environ.get("GITHUB_REF_NAME", "main"))
INCLUDE_APP_VIEW = os.environ.get("DASHBOARD_INCLUDE_APP_VIEW", "1").strip().lower() not in {"0", "false", "no"}


def rows_from_sheet(wb, sheet_name):
    ws = wb[sheet_name]
    headers = [str(cell.value or "") for cell in ws[1]]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None and str(v).strip() for v in values):
            continue
        rows.append({h: ("" if v is None else v) for h, v in zip(headers, values)})
    return rows


wb = openpyxl.load_workbook(BASE, data_only=True)
base = {
    "sites": rows_from_sheet(wb, "Resumen por sede"),
    "details": rows_from_sheet(wb, "Detalle por fecha"),
    "files": rows_from_sheet(wb, "Archivos SharePoint"),
    "calendar_validation": rows_from_sheet(wb, "Validación calendario"),
}
pillars = json.loads(PILLARS.read_text(encoding="utf-8"))
if APP_UPLOADS.exists():
    app_uploads = json.loads(APP_UPLOADS.read_text(encoding="utf-8"))
else:
    app_uploads = {
        "enabled": False,
        "message": "No se encontro validacion del aplicativo. Configure APP_AUTOGESTIONES_CSV_PATH para activar esta vista.",
        "source": "",
        "summary": {},
        "rows": [],
        "detail": [],
        "unmatched": [],
    }


def rollup_files(checklists):
    files = {}
    for checklist in checklists:
        archivo = checklist.get("archivo", "")
        if not archivo:
            continue
        row = files.setdefault(
            archivo,
            {
                "checklists_pdf": 0,
                "horas_pdf": 0.0,
                "autogestiones_pdf": 0,
                "pilares": {},
            },
        )
        row["checklists_pdf"] += 1
        row["horas_pdf"] += float(checklist.get("horas_calculadas") or 0)
        row["autogestiones_pdf"] += int(checklist.get("autogestiones_pdf") or 0)
        for pilar, cantidad in (checklist.get("pilares") or {}).items():
            row["pilares"][pilar] = row["pilares"].get(pilar, 0) + int(cantidad or 0)

    for row in files.values():
        row["horas_pdf"] = round(row["horas_pdf"], 2)
        row["pilares_pdf"] = " | ".join(
            f"{pilar}: {cantidad}" for pilar, cantidad in sorted(row["pilares"].items())
        ) or "Sin pilares"
        del row["pilares"]
    return files


file_rollup = rollup_files(pillars.get("checklists", []))
for file_row in base["files"]:
    file_row.update(
        file_rollup.get(
            file_row.get("archivo", ""),
            {
                "checklists_pdf": 0,
                "horas_pdf": 0,
                "autogestiones_pdf": 0,
                "pilares_pdf": "Sin lectura PDF",
            },
        )
    )

data = {**base, **pillars, "app_uploads": app_uploads}
if not INCLUDE_APP_VIEW:
    data["app_uploads"] = {
        "enabled": False,
        "message": "Vista local no incluida en la publicacion de GitHub.",
        "source": "",
        "summary": {},
        "rows": [],
        "detail": [],
        "unmatched": [],
    }
payload = json.dumps(data, ensure_ascii=False)
app_tab_html = '<button class="tab" data-tab="app">Falta subir aplicativo</button>' if INCLUDE_APP_VIEW else ""
include_app_js = "true" if INCLUDE_APP_VIEW else "false"
panel_names = ["sites", "pillars", "checks", "alerts", "files", "calendar"]
if INCLUDE_APP_VIEW:
    panel_names.append("app")
panel_names.append("days")
panel_names_json = json.dumps(panel_names, ensure_ascii=False)

html = f"""<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Dashboard Soportes, Pilares y Horas {CARPETA_MES} {ANIO}</title>
<style>
:root {{
  --bg:#f5f7fa; --panel:#fff; --ink:#17202a; --muted:#637083; --line:#dce3ec;
  --blue:#071936; --red:#c63d3d; --green:#1f8a5b; --amber:#b7791f; --soft:#eef3f8;
  --brand:#7ee000;
  --shadow:0 10px 24px rgba(20,31,45,.08);
}}
*{{box-sizing:border-box}}
body{{margin:0;background:var(--bg);color:var(--ink);font-family:Arial,Helvetica,sans-serif;font-size:14px;letter-spacing:0;position:relative;min-height:100vh}}
body::before{{content:"Realizado por Bryan Martinez";position:fixed;left:50%;top:56%;transform:translate(-50%,-50%) rotate(-22deg);font-size:clamp(34px,6vw,86px);font-weight:900;letter-spacing:.04em;color:#071936;opacity:.045;white-space:nowrap;pointer-events:none;z-index:0}}
header{{background:linear-gradient(135deg,#071936 0%,#0a2144 58%,#0f3159 100%);color:#fff;border-bottom:4px solid var(--brand);padding:0;position:sticky;top:0;z-index:4;box-shadow:0 10px 24px rgba(7,25,54,.22)}}
.wrap{{max-width:1480px;margin:0 auto}}
.hero{{display:block;min-height:116px;position:relative;overflow:hidden}}
.hero::after{{content:"";position:absolute;right:-80px;top:-70px;width:330px;height:250px;border:28px solid rgba(126,224,0,.28);transform:rotate(17deg);border-radius:22px;pointer-events:none}}
.hero-copy{{padding:18px 22px 16px;position:relative;z-index:1}}
.hero-brand{{display:none}}
.top{{display:flex;align-items:center;justify-content:space-between;gap:14px}}
h1{{margin:0;font-size:24px;line-height:1.12}} .sub{{display:block;color:#d9e5f0;margin-top:7px;max-width:820px;line-height:1.45}}
.brand-accent{{color:#a9d800}}
.filters{{display:grid;grid-template-columns:minmax(260px,1fr) 170px 170px 170px;gap:8px;margin-top:10px}}
input,select,button{{height:34px;border:1px solid var(--line);border-radius:6px;background:#fff;color:var(--ink);padding:0 10px;font:inherit;min-width:0}}
button{{cursor:pointer;background:#113c5f;color:#fff;border-color:#113c5f;font-weight:700}}
button:disabled{{opacity:.65;cursor:wait}}
.actions{{display:flex;gap:8px;align-items:center;flex-wrap:wrap;justify-content:flex-end}}
.session-heading{{margin:6px 0 14px;padding:4px 2px 0}}
.session-eyebrow{{display:inline-flex;align-items:center;gap:8px;color:#0f3159;text-transform:uppercase;font-size:11px;font-weight:900;letter-spacing:.12em}}
.session-eyebrow::before{{content:"";display:inline-block;width:28px;height:4px;border-radius:999px;background:var(--brand)}}
.session-heading h2{{font-size:22px;margin:6px 0 4px;color:#071936}}
.session-heading p{{margin:0;color:var(--muted);max-width:930px;line-height:1.45}}
.session-card{{border-top:5px solid #113c5f}}
.session-card.audit{{border-top-color:var(--brand)}}
.refresh-status{{min-height:0;margin-top:4px;color:#dbe5ef;font-size:12px;text-align:right}}
.refresh-status.bad-text{{color:#9c2d2d}}
main{{max-width:1480px;margin:0 auto;padding:18px 24px 36px;position:relative;z-index:1}}
.kpis{{display:grid;grid-template-columns:repeat(8,minmax(120px,1fr));gap:12px;margin-bottom:16px}}
.metric{{background:rgba(255,255,255,.95);border:1px solid var(--line);border-radius:8px;padding:12px 13px;box-shadow:var(--shadow);min-height:82px;backdrop-filter:blur(2px)}}
.metric span{{display:block;color:var(--muted);font-size:12px;margin-bottom:8px}} .metric strong{{font-size:24px;line-height:1}}
.grid{{display:grid;grid-template-columns:1fr 1fr;gap:16px;align-items:start}}
section{{background:rgba(255,255,255,.95);border:1px solid var(--line);border-radius:8px;padding:16px;box-shadow:var(--shadow);margin-bottom:16px;backdrop-filter:blur(2px)}}
h2{{font-size:16px;margin:0 0 12px}}
.tabs{{display:flex;gap:8px;flex-wrap:wrap;margin:0 0 12px}} .tab{{background:#fff;color:var(--ink);border-color:var(--line);padding:0 12px}} .tab.active{{background:#113c5f;color:#fff;border-color:#113c5f}}
.table-wrap{{overflow:auto;max-height:590px;border:1px solid var(--line);border-radius:8px}} table{{width:100%;border-collapse:collapse;min-width:1050px;background:#fff}}
th,td{{text-align:left;padding:9px 10px;border-bottom:1px solid var(--line);vertical-align:top}} th{{position:sticky;top:0;background:var(--soft);z-index:1;font-size:12px;text-transform:uppercase;color:#465568}}
.group-head th{{top:0;background:#d7e6f3;color:#16354f;text-align:center;font-size:12px;letter-spacing:0}}
.group-head + tr th{{top:35px}}
.panel-note{{padding:10px 12px;border-bottom:1px solid var(--line);background:#fbfdff;color:var(--muted);font-size:13px}}
.stack{{display:grid;gap:4px;min-width:120px}} .stack strong{{font-size:17px;line-height:1;color:#17202a}} .stack span{{color:var(--muted);font-size:12px;line-height:1.25}}
.days{{max-width:260px;white-space:normal;word-break:break-word}} .warn-text{{color:#8a5a08}} .bad-text{{color:#9c2d2d}}
td.num,th.num{{text-align:right;font-variant-numeric:tabular-nums}} tr:hover td{{background:#f8fbfe}}
.pill{{display:inline-flex;min-width:82px;justify-content:center;border-radius:999px;padding:4px 8px;font-size:12px;font-weight:700}}
.ok{{background:#ddf5e9;color:#0f6845}} .bad{{background:#fde2e2;color:#9c2d2d}} .warn{{background:#fff0d1;color:#8a5a08}}
.bars{{display:grid;gap:9px}} .bar{{display:grid;grid-template-columns:minmax(160px,1fr) 3fr 58px;gap:10px;align-items:center}}
.label{{white-space:nowrap;overflow:hidden;text-overflow:ellipsis}} .track{{height:14px;background:#edf1f6;border-radius:4px;overflow:hidden}} .fill{{height:100%;background:var(--red)}} .fill.green{{background:var(--green)}} .val{{text-align:right;color:var(--muted);font-variant-numeric:tabular-nums}}
.day-tools{{display:flex;gap:8px;align-items:center;flex-wrap:wrap;padding:12px;border-bottom:1px solid var(--line);background:#fbfdff}}
.day-buttons{{display:flex;gap:6px;flex-wrap:wrap}}
.day-btn{{width:36px;height:32px;padding:0;background:#fff;color:var(--ink);border-color:var(--line);font-weight:700}}
.day-btn.active{{background:#113c5f;color:#fff;border-color:#113c5f}}
.ghost-btn{{background:#fff;color:#113c5f;border-color:#bfd0df}}
.ghost-btn.active{{background:#113c5f;color:#fff;border-color:#113c5f}}
.day-kpis{{display:grid;grid-template-columns:repeat(5,minmax(120px,1fr));gap:10px;padding:12px;border-bottom:1px solid var(--line);background:#fff}}
.day-kpi{{border:1px solid var(--line);border-radius:6px;padding:10px;background:#f8fbfe;min-height:66px}}
.day-kpi span{{display:block;color:var(--muted);font-size:12px;margin-bottom:6px}} .day-kpi strong{{font-size:22px}}
.audit-tools{{display:flex;gap:8px;align-items:center;flex-wrap:wrap;padding:12px;border-bottom:1px solid var(--line);background:#fbfdff}}
.audit-status{{height:32px;border-radius:999px;background:#fff;color:#113c5f;border:1px solid #bfd0df;padding:0 12px;font-weight:700}}
.audit-status.active{{background:#113c5f;color:#fff;border-color:#113c5f}}
.audit-status.danger.active{{background:#9c2d2d;border-color:#9c2d2d}}
.audit-status.warn.active{{background:#8a5a08;border-color:#8a5a08}}
.audit-table{{min-width:980px}}
.site-cell strong{{display:block;font-size:13px;line-height:1.25;max-width:330px}}
.site-cell span{{display:block;margin-top:3px;color:var(--muted);font-size:12px}}
.count-cell{{text-align:center;font-variant-numeric:tabular-nums}}
.count-cell strong{{display:block;font-size:22px;line-height:1;color:#17202a}}
.count-cell span{{display:block;margin-top:4px;color:var(--muted);font-size:11px;text-transform:uppercase;letter-spacing:.02em}}
.day-chip-list{{display:flex;gap:4px;flex-wrap:wrap;max-width:360px}}
.day-chip{{display:inline-flex;align-items:center;justify-content:center;min-width:26px;height:24px;border-radius:999px;padding:0 7px;font-size:12px;font-weight:800;border:1px solid var(--line);background:#f8fbfe;color:#17202a}}
.day-chip.missing{{background:#fde2e2;border-color:#f5bcbc;color:#9c2d2d}}
.day-chip.done{{background:#ddf5e9;border-color:#b9e8d0;color:#0f6845}}
.day-chip.outside{{background:#fff0d1;border-color:#f1d38b;color:#8a5a08}}
.muted-line{{display:block;color:var(--muted);font-size:12px;margin-top:5px}}
.audit-empty{{padding:18px;color:var(--muted)}}
.hidden{{display:none}} .small{{color:var(--muted);font-size:12px}}
a{{color:var(--blue);text-decoration:none}} a:hover{{text-decoration:underline}}
@media(max-width:1120px){{.filters,.kpis,.grid,.day-kpis{{grid-template-columns:1fr 1fr}}}} @media(max-width:740px){{header{{position:static}}.hero-copy{{padding:14px}}main{{padding:16px}}.top{{align-items:flex-start;flex-direction:column}}.filters,.kpis,.grid,.day-kpis{{grid-template-columns:1fr}}h1{{font-size:20px}}.bar{{grid-template-columns:1fr}}}}
</style>
</head>
<body>
<header>
  <div class="hero wrap">
    <div class="hero-copy">
    <div class="top">
      <div>
        <h1>Centro de control documental - {CARPETA_MES} {ANIO}</h1>
        <div class="sub">Seguimientos separados para soportes, horas, pilares, alertas y trazabilidad mensual.</div>
      </div>
      <div class="actions">
      
        <button id="csvBtn" type="button">Descargar CSV</button>
      </div>
    </div>
    <div id="refreshStatus" class="refresh-status"></div>
    <div class="filters">
      <input id="search" type="search" placeholder="Buscar sede, pilar, archivo o técnico">
      <select id="status"><option value="all">Todos los soportes</option><option value="Completo">Soportes completos</option><option value="Incompleto">Soportes incompletos</option></select>
      <select id="ues"><option value="all">Todas las UES</option></select>
      <select id="onlyAlerts"><option value="all">Todas las filas</option><option value="alerts">Solo alertas</option></select>
    </div>
    </div>
  </div>
</header>
<main>
  <div class="session-heading">
    <span class="session-eyebrow">Sesión 1 · Seguimiento documental</span>
    <h2>Resumen ejecutivo de soportes</h2>
    <p>Indicadores principales para revisar cobertura, faltantes de soporte y diferencias de horas antes de entrar al detalle operativo.</p>
  </div>
  <div class="kpis">
    <div class="metric"><span>PDFs leídos</span><strong id="kPdfs">0</strong></div>
    <div class="metric"><span>Checklists leídos</span><strong id="kChecks">0</strong></div>
    <div class="metric"><span>Horas calculadas</span><strong id="kHours">0</strong></div>
    <div class="metric"><span>Autogestiones PDF</span><strong id="kAuto">0</strong></div>
    <div class="metric"><span>Alertas clasificadas</span><strong id="kAlerts">0</strong></div>
    <div class="metric"><span>Archivos SharePoint</span><strong id="kFiles">0</strong></div>
    <div class="metric"><span>Faltantes soporte</span><strong id="kMissing">0</strong></div>
    <div class="metric"><span>Cobertura soporte</span><strong id="kRate">0%</strong></div>
  </div>

  <div class="grid">
    <section class="session-card"><h2>Mayores faltantes de soporte</h2><div id="missingBars" class="bars"></div></section>
    <section class="session-card"><h2>Diferencias de horas en sedes completas</h2><div id="hourBars" class="bars"></div><div class="small">Las sedes incompletas no se usan para concluir diferencia mensual de horas.</div></section>
  </div>

  <div class="session-heading">
    <span class="session-eyebrow">Sesión 2 · Auditoría por fuente</span>
    <h2>Validaciones detalladas</h2>
    <p>Vistas de revisión para comparar cronograma, soportes PDF, SharePoint, pilares y alertas documentales.</p>
  </div>
  <section class="session-card audit">
    <div class="tabs">
      <button class="tab active" data-tab="sites">Soportes por sede</button>
      <button class="tab" data-tab="pillars">Horas y pilares</button>
      <button class="tab" data-tab="checks">Detalle checklists</button>
      <button class="tab" data-tab="alerts">Alertas</button>
      <button class="tab" data-tab="files">Archivos</button>
      <button class="tab" data-tab="calendar">Esperado vs Cronograma vs SharePoint</button>
      {app_tab_html}
      <button class="tab" data-tab="days">Cálculo por días</button>
    </div>

    <div id="panel-sites" class="table-wrap"><table><thead><tr><th>Sede</th><th>UES</th><th>Ubicación</th><th class="num">Digitados</th><th class="num">Cubiertos</th><th class="num">Falta soporte</th><th>Días digitados sin soporte</th><th>Estado</th></tr></thead><tbody id="rowsSites"></tbody></table></div>
    <div id="panel-pillars" class="table-wrap hidden"><table><thead><tr><th>Sede</th><th>Estado soportes</th><th class="num">Checklists PDF</th><th class="num">Horas PDF</th><th class="num">Horas cronog</th><th class="num">Dif horas</th><th class="num">Autog. PDF</th><th class="num">Autog. cronog</th><th class="num">Dif autog.</th><th>Pilares detectados</th></tr></thead><tbody id="rowsPillars"></tbody></table></div>
    <div id="panel-checks" class="table-wrap hidden"><table><thead><tr><th>No. autogestión</th><th>Sede</th><th>Fecha</th><th>Horario</th><th>Técnicos detectados</th><th class="num">Cant.</th><th class="num">Horas calc.</th><th class="num">Autog.</th><th>Pilares</th><th>Archivo</th></tr></thead><tbody id="rowsChecks"></tbody></table></div>
    <div id="panel-alerts" class="table-wrap hidden"><table><thead><tr><th>Categoría</th><th>Severidad</th><th>Alerta</th><th>Sede</th><th>No. autogestión</th><th>Fecha</th><th>Campo</th><th class="num">PDF</th><th class="num">Cronograma</th><th class="num">Diferencia</th><th class="num">Actividades PDF</th><th class="num">Con cobro PDF</th><th class="num">Sin cobro PDF</th><th>Detalle</th><th>Archivo</th></tr></thead><tbody id="rowsAlerts"></tbody></table></div>
    <div id="panel-files" class="table-wrap hidden"><table><thead><tr><th>Archivo</th><th>Carpeta</th><th>Sede asociada</th><th>Días leídos en PDF</th><th>Fechas leídas en PDF</th><th class="num">Checklists PDF</th><th class="num">Horas PDF</th><th class="num">Autog. PDF</th><th>Consolidado por pilar</th><th>Fuente días</th><th class="num">Confianza</th><th>Fecha carga</th></tr></thead><tbody id="rowsFiles"></tbody></table></div>
    <div id="panel-calendar" class="table-wrap hidden">
      <div class="panel-note">Esta vista compara tres fuentes: lo esperado según la programación diaria de General MAYO, lo digitado en Cronog y lo cargado en SharePoint.</div>
      <table><thead>
        <tr class="group-head"><th colspan="3">Sede</th><th>Esperado según General</th><th>Digitado en Cronograma</th><th>Cargado en SharePoint</th><th colspan="2">Pendientes</th><th>Resultado</th></tr>
        <tr><th>Sede</th><th>UES</th><th>Rango</th><th>Checklist esperados</th><th>Fechas digitadas</th><th>Soportes cargados</th><th>Por diligenciar</th><th>Por cargar</th><th>Estado / alertas</th></tr>
      </thead><tbody id="rowsCalendar"></tbody></table>
    </div>
    <div id="panel-app" class="table-wrap hidden">
      <div class="panel-note">Esta vista compara los dias esperados segun General/cuadritos contra el CSV Master Autogestiones para identificar lo que falta por subir al aplicativo.</div>
      <div class="day-kpis">
        <div class="day-kpi"><span>Sedes revisadas</span><strong id="aSites">0</strong></div>
        <div class="day-kpi"><span>Esperados General</span><strong id="aExpected">0</strong></div>
        <div class="day-kpi"><span>Subidos aplicativo</span><strong id="aUploaded">0</strong></div>
        <div class="day-kpi"><span>Faltan aplicativo</span><strong id="aMissing">0</strong></div>
        <div class="day-kpi"><span>Cobertura aplicativo</span><strong id="aRate">0%</strong></div>
      </div>
      <div class="audit-tools">
        <button class="audit-status active danger" data-app-filter="missing" type="button">Pendientes</button>
        <button class="audit-status" data-app-filter="all" type="button">Todas</button>
        <button class="audit-status" data-app-filter="complete" type="button">Completas</button>
        <button class="audit-status warn" data-app-filter="outside" type="button">Fuera programacion</button>
      </div>
      <table class="audit-table"><thead>
        <tr><th>Sede</th><th class="num">Esperado</th><th class="num">Subido</th><th class="num">Falta</th><th>Dias pendientes</th><th>Dias subidos</th><th>Estado</th></tr>
      </thead><tbody id="rowsApp"></tbody></table>
    </div>
    <div id="panel-days" class="table-wrap hidden">
      <div class="panel-note">Selecciona uno o varios días para calcular únicamente los soportes cargados cuyas fechas fueron leídas dentro del PDF. Sin selección se toman todos los días disponibles.</div>
      <div class="day-tools">
        <button id="allDaysBtn" class="ghost-btn" type="button">Todos</button>
        <button id="clearDaysBtn" class="ghost-btn" type="button">Limpiar</button>
        <div id="dayButtons" class="day-buttons"></div>
      </div>
      <div class="day-kpis">
        <div class="day-kpi"><span>Días incluidos</span><strong id="dDays">0</strong></div>
        <div class="day-kpi"><span>Archivos con soporte</span><strong id="dFiles">0</strong></div>
        <div class="day-kpi"><span>Sedes con soporte</span><strong id="dSites">0</strong></div>
        <div class="day-kpi"><span>Checklists</span><strong id="dChecks">0</strong></div>
        <div class="day-kpi"><span>Horas PDF</span><strong id="dHours">0</strong></div>
      </div>
      <table><thead><tr><th>Archivo</th><th>Sede</th><th>Días incluidos</th><th class="num">Checklists</th><th class="num">Horas PDF</th><th class="num">Autog. PDF</th><th>Consolidado por pilar</th></tr></thead><tbody id="rowsDays"></tbody></table>
    </div>
  </section>
</main>
<script>
const DATA = {payload};
const INCLUDE_APP_VIEW = {include_app_js};
const REFRESH_ENDPOINT = "{REFRESH_ENDPOINT}";
const REFRESH_CONFIG = {{anio:"{ANIO}", mes:"{CARPETA_MES.split('-', 1)[0]}"}};
const GITHUB_REFRESH = {{
  repository: "{GITHUB_REPOSITORY}",
  workflow: "{GITHUB_WORKFLOW_FILE}",
  branch: "{GITHUB_BRANCH}",
  anio: "{ANIO}",
  mes: "{CARPETA_MES.split('-', 1)[0]}"
}};
const state = {{tab:"sites", search:"", status:"all", ues:"all", onlyAlerts:"all", appFilter:"missing", selectedDays:new Set()}};
const $ = id => document.getElementById(id);
const fmt = new Intl.NumberFormat("es-CO", {{maximumFractionDigits:1}});
const esc = v => String(v ?? "").replace(/[&<>"']/g, c => ({{"&":"&amp;","<":"&lt;",">":"&gt;",'"':"&quot;","'":"&#39;"}}[c]));
const num = v => Number(String(v ?? 0).replace(",", ".")) || 0;
const norm = v => String(v ?? "").toLowerCase().normalize("NFD").replace(/[\\u0300-\\u036f]/g,"");
const pCodes = ["INS-ELE","SEÑA","OBR-CIV","CARP-MET","CHA-LLAV","PINT MTTO","HID-SAN","CUBI","APY-LOG","APY-MTTO","CARP-MAD","ADE-OFI","PERSIA"];

function includesQ(row) {{ const q=norm(state.search); return !q || norm(Object.values(row).join(" ")).includes(q); }}
function siteOk(row) {{ return includesQ(row) && (state.status==="all" || row.estado===state.status) && (state.ues==="all" || row.ues===state.ues); }}
function pillarOk(row) {{ return includesQ(row) && (state.status==="all" || row.estado_soportes===state.status) && (state.onlyAlerts==="all" || (row.estado_soportes==="Completo" && (Math.abs(num(row.dif_horas_pdf_vs_cronograma))>.25 || Math.abs(num(row.dif_autogestiones_pdf_vs_cronograma))>.25))); }}
function badge(value) {{ return `<span class="pill ${{value==="Completo"?"ok":value==="Cargado"?"ok":value==="Incompleto"?"bad":"warn"}}">${{esc(value || "Parcial")}}</span>`; }}
function severityBadge(value) {{ return `<span class="pill ${{value==="Crítica"?"bad":value==="Revisar"?"warn":"ok"}}">${{esc(value || "Info")}}</span>`; }}
function pillarSummary(row) {{ return pCodes.filter(c => num(row["pdf_"+c]) || num(row["cron_"+c])).map(c => `${{c}} PDF:${{fmt.format(num(row["pdf_"+c]))}} / Cron:${{fmt.format(num(row["cron_"+c]))}}`).join("<br>"); }}
function checkPillars(row) {{ return Object.entries(row.pilares || {{}}).map(([k,v]) => `${{k}}:${{v}}`).join(", "); }}
function stack(count, lines, cls="") {{ return `<div class="stack ${{cls}}"><strong>${{fmt.format(num(count))}}</strong>${{lines.filter(Boolean).map(v=>`<span class="days">${{esc(v)}}</span>`).join("")}}</div>`; }}
function dayChips(value, cls="") {{
  const days = String(value || "").match(/\\d+/g) || [];
  return days.length ? `<div class="day-chip-list">${{days.map(d=>`<span class="day-chip ${{cls}}">${{esc(d)}}</span>`).join("")}}</div>` : `<span class="small">Sin dias</span>`;
}}
function compactAutos(value) {{
  const autos = String(value || "").split(",").map(v=>v.trim()).filter(Boolean);
  if (!autos.length) return "Sin autogestiones";
  const head = autos.slice(0, 5).join(", ");
  return autos.length > 5 ? `${{head}} +${{autos.length - 5}} mas` : head;
}}
function checklistDay(row) {{ const m=String(row.fecha || "").match(/-(\\d{{2}})$/); return m ? Number(m[1]) : 0; }}
function availableDays() {{ return [...new Set(DATA.checklists.map(checklistDay).filter(Boolean))].sort((a,b)=>a-b); }}
function activeDays() {{ const days=availableDays(); return state.selectedDays.size ? [...state.selectedDays].sort((a,b)=>a-b) : days; }}
function fileUrl(archivo) {{ const found=DATA.files.find(f=>f.archivo===archivo); return found ? found.url : ""; }}
function siteRowFor(name) {{
  const n=norm(name);
  return DATA.sites.find(s=>norm(s.sede)===n) || DATA.sites.find(s=>n.includes(norm(s.sede)) || norm(s.sede).includes(n)) || null;
}}
function sitePillarRowFor(name) {{
  const n=norm(name);
  return DATA.site_rows.find(s=>norm(s.sede)===n) || DATA.site_rows.find(s=>n.includes(norm(s.sede)) || norm(s.sede).includes(n)) || {{}};
}}
function checkMatchesDayFilters(row) {{
  const q=norm(state.search);
  const site=siteRowFor(row.sede_pdf);
  const text=norm([row.autogestion,row.sede_pdf,row.fecha,row.tecnicos_nombres,row.archivo,checkPillars(row),JSON.stringify(row.pilares||{{}})].join(" "));
  return (!q || text.includes(q)) && (state.ues==="all" || (site && site.ues===state.ues));
}}
function dayChecks() {{
  const days=new Set(activeDays());
  return DATA.checklists.filter(r=>days.has(checklistDay(r)) && checkMatchesDayFilters(r));
}}
function dayRows() {{
  const rows=new Map();
  for (const chk of dayChecks()) {{
    const key=chk.archivo || "Sin archivo";
    if (!rows.has(key)) rows.set(key, {{archivo:key, url:fileUrl(key), sede:chk.sede_pdf || "", dias:new Set(), checklists:0, horas_pdf:0, autogestiones_pdf:0, pilares:{{}}}});
    const row=rows.get(key);
    row.dias.add(checklistDay(chk));
    row.checklists += 1;
    row.horas_pdf += num(chk.horas_calculadas);
    row.autogestiones_pdf += num(chk.autogestiones_pdf);
    for (const [pilar,cantidad] of Object.entries(chk.pilares || {{}})) row.pilares[pilar]=(row.pilares[pilar]||0)+num(cantidad);
  }}
  return [...rows.values()].map(r=>({{
    ...r,
    dias_texto:[...r.dias].sort((a,b)=>a-b).join(","),
    horas_pdf:Number(r.horas_pdf.toFixed(2)),
    pilares_pdf:Object.entries(r.pilares).sort((a,b)=>a[0].localeCompare(b[0])).map(([k,v])=>`${{k}}: ${{fmt.format(v)}}`).join(" | ") || "Sin pilares"
  }})).sort((a,b)=>String(a.sede).localeCompare(String(b.sede)) || String(a.archivo).localeCompare(String(b.archivo)));
}}
function renderDayButtons() {{
  const allMode=!state.selectedDays.size;
  $("allDaysBtn").classList.toggle("active", allMode);
  $("dayButtons").innerHTML = availableDays().map(day=>`<button type="button" class="day-btn ${{state.selectedDays.has(day) ? "active" : ""}}" data-day="${{day}}">${{day}}</button>`).join("");
  document.querySelectorAll(".day-btn").forEach(btn=>btn.addEventListener("click",()=>{{ const day=Number(btn.dataset.day); state.selectedDays.has(day) ? state.selectedDays.delete(day) : state.selectedDays.add(day); render(); }}));
}}
function renderDayCalc() {{
  renderDayButtons();
  const rows=dayRows();
  const checks=dayChecks();
  $("dDays").textContent = activeDays().join(",") || "0";
  $("dFiles").textContent = fmt.format(rows.length);
  $("dSites").textContent = fmt.format(new Set(checks.map(r=>r.sede_pdf)).size);
  $("dChecks").textContent = fmt.format(checks.length);
  $("dHours").textContent = fmt.format(checks.reduce((s,r)=>s+num(r.horas_calculadas),0));
  $("rowsDays").innerHTML = rows.map(r=>`<tr><td>${{r.url ? `<a target="_blank" rel="noopener" href="${{esc(r.url)}}">${{esc(r.archivo)}}</a>` : esc(r.archivo)}}</td><td>${{esc(r.sede)}}</td><td>${{esc(r.dias_texto)}}</td><td class="num">${{fmt.format(num(r.checklists))}}</td><td class="num">${{fmt.format(num(r.horas_pdf))}}</td><td class="num">${{fmt.format(num(r.autogestiones_pdf))}}</td><td class="days">${{esc(r.pilares_pdf)}}</td></tr>`).join("") || `<tr><td colspan="7">No hay soportes cargados para la selección actual.</td></tr>`;
}}

function fillFilters() {{
  const ues = [...new Set(DATA.sites.map(r=>r.ues).filter(Boolean))].sort();
  $("ues").innerHTML += ues.map(v=>`<option value="${{esc(v)}}">${{esc(v)}}</option>`).join("");
}}
function renderKpis() {{
  const sites = DATA.sites.filter(siteOk);
  const typed = sites.reduce((s,r)=>s+num(r.cantidad_digitada),0);
  const covered = sites.reduce((s,r)=>s+num(r.cantidad_cubierta),0);
  const missing = sites.reduce((s,r)=>s+num(r.cantidad_faltante),0);
  $("kPdfs").textContent = fmt.format(DATA.summary.pdfs_leidos);
  $("kChecks").textContent = fmt.format(DATA.summary.checklists_leidos);
  $("kHours").textContent = fmt.format(DATA.summary.horas_calculadas);
  $("kAuto").textContent = fmt.format(DATA.summary.autogestiones_detectadas);
  $("kAlerts").textContent = fmt.format(DATA.summary.alertas);
  $("kFiles").textContent = fmt.format(DATA.files.length);
  $("kMissing").textContent = fmt.format(missing);
  $("kRate").textContent = typed ? Math.round(covered/typed*100)+"%" : "0%";
}}
function bars() {{
  const sites = DATA.sites.filter(siteOk).sort((a,b)=>num(b.cantidad_faltante)-num(a.cantidad_faltante)).slice(0,12);
  const maxM = Math.max(1, ...sites.map(r=>num(r.cantidad_faltante)));
  $("missingBars").innerHTML = sites.map(r=>`<div class="bar"><div class="label" title="${{esc(r.sede)}}">${{esc(r.sede)}}</div><div class="track"><div class="fill" style="width:${{Math.max(3,num(r.cantidad_faltante)/maxM*100)}}%"></div></div><div class="val">${{fmt.format(num(r.cantidad_faltante))}}</div></div>`).join("");
  const complete = DATA.site_rows.filter(r=>r.estado_soportes==="Completo" && includesQ(r)).sort((a,b)=>Math.abs(num(b.dif_horas_pdf_vs_cronograma))-Math.abs(num(a.dif_horas_pdf_vs_cronograma))).slice(0,12);
  const maxH = Math.max(1, ...complete.map(r=>Math.abs(num(r.dif_horas_pdf_vs_cronograma))));
  $("hourBars").innerHTML = complete.map(r=>`<div class="bar"><div class="label" title="${{esc(r.sede)}}">${{esc(r.sede)}}</div><div class="track"><div class="fill green" style="width:${{Math.max(3,Math.abs(num(r.dif_horas_pdf_vs_cronograma))/maxH*100)}}%"></div></div><div class="val">${{fmt.format(num(r.dif_horas_pdf_vs_cronograma))}}</div></div>`).join("") || "<div class='small'>Sin sedes completas con los filtros actuales.</div>";
}}
function appRows() {{
  const app = DATA.app_uploads || {{}};
  const rows = app.rows || [];
  return rows.filter(r =>
    includesQ(r) &&
    (state.ues==="all" || r.ues===state.ues) &&
    (state.status==="all" || r.estado===state.status) &&
    (state.onlyAlerts==="all" || num(r.cantidad_falta_subir_aplicativo)>0) &&
    (state.appFilter==="all" ||
      (state.appFilter==="missing" && num(r.cantidad_falta_subir_aplicativo)>0) ||
      (state.appFilter==="complete" && num(r.cantidad_falta_subir_aplicativo)===0) ||
      (state.appFilter==="outside" && num(r.cantidad_fuera_programacion_aplicativo)>0))
  ).sort((a,b)=>num(b.cantidad_falta_subir_aplicativo)-num(a.cantidad_falta_subir_aplicativo) || String(a.sede).localeCompare(String(b.sede)));
}}
function renderAppUploads() {{
  if (!INCLUDE_APP_VIEW || !$("rowsApp")) return;
  const app = DATA.app_uploads || {{}};
  const summary = app.summary || {{}};
  $("aSites").textContent = fmt.format(summary.sedes_revisadas || 0);
  $("aExpected").textContent = fmt.format(summary.checklists_esperados || 0);
  $("aUploaded").textContent = fmt.format(summary.cargados_aplicativo || 0);
  $("aMissing").textContent = fmt.format(summary.faltan_aplicativo || 0);
  $("aRate").textContent = (summary.cobertura_aplicativo || 0) + "%";
  if (!app.enabled) {{
    $("rowsApp").innerHTML = `<tr><td colspan="7" class="audit-empty"><strong>Validacion no disponible.</strong><br><span class="small">${{esc(app.message || "No se encontro el CSV Master Autogestiones.")}}</span></td></tr>`;
    return;
  }}
  const rows = appRows();
  $("rowsApp").innerHTML = rows.map(r=>`<tr>
    <td class="site-cell"><strong>${{esc(r.sede)}}</strong><span>${{esc(r.ues || "Sin UES")}} · ${{esc(r.fecha_inicio)}}-${{esc(r.fecha_fin)}}</span></td>
    <td class="count-cell"><strong>${{fmt.format(num(r.cantidad_esperada_general))}}</strong><span>General</span></td>
    <td class="count-cell"><strong>${{fmt.format(num(r.cantidad_subida_aplicativo))}}</strong><span>CSV</span></td>
    <td class="count-cell ${{num(r.cantidad_falta_subir_aplicativo) ? "bad-text" : ""}}"><strong>${{fmt.format(num(r.cantidad_falta_subir_aplicativo))}}</strong><span>Pendiente</span></td>
    <td>${{dayChips(r.dias_falta_subir_aplicativo, "missing")}}<span class="muted-line">${{num(r.cantidad_falta_subir_aplicativo) ? "Requiere cargar en aplicativo" : "Sin pendientes"}}</span></td>
    <td>${{dayChips(r.dias_subidos_aplicativo, "done")}}<span class="muted-line">${{esc(compactAutos(r.autogestiones_csv))}}</span></td>
    <td>${{badge(r.estado)}}<span class="muted-line">${{num(r.cantidad_fuera_programacion_aplicativo) ? "Fuera: " + esc(r.dias_fuera_programacion_aplicativo) : esc(r.alerta || "")}}</span></td>
  </tr>`).join("") || `<tr><td colspan="7" class="audit-empty">Sin registros con los filtros actuales.</td></tr>`;
}}
function renderTables() {{
  const sites = DATA.sites.filter(siteOk).sort((a,b)=>num(b.cantidad_faltante)-num(a.cantidad_faltante));
  $("rowsSites").innerHTML = sites.map(r=>`<tr><td>${{esc(r.sede)}}</td><td>${{esc(r.ues)}}</td><td>${{esc(r.ubicacion)}}</td><td class="num">${{fmt.format(num(r.cantidad_digitada))}}</td><td class="num">${{fmt.format(num(r.cantidad_cubierta))}}</td><td class="num">${{fmt.format(num(r.cantidad_faltante))}}</td><td>${{esc(r.dias_faltantes || "Sin faltantes")}}</td><td>${{badge(r.estado)}}</td></tr>`).join("");
  const ps = DATA.site_rows.filter(pillarOk).sort((a,b)=>String(a.sede).localeCompare(String(b.sede)));
  $("rowsPillars").innerHTML = ps.map(r=>`<tr><td>${{esc(r.sede)}}</td><td>${{badge(r.estado_soportes)}}</td><td class="num">${{fmt.format(num(r.checklists_pdf))}}</td><td class="num">${{fmt.format(num(r.horas_calculadas_pdf))}}</td><td class="num">${{fmt.format(num(r.horas_reales_cronograma))}}</td><td class="num">${{fmt.format(num(r.dif_horas_pdf_vs_cronograma))}}</td><td class="num">${{fmt.format(num(r.autogestiones_pdf))}}</td><td class="num">${{fmt.format(num(r.total_autogestiones_cronograma))}}</td><td class="num">${{fmt.format(num(r.dif_autogestiones_pdf_vs_cronograma))}}</td><td>${{pillarSummary(r)}}</td></tr>`).join("");
  const checks = DATA.checklists.filter(includesQ).sort((a,b)=>String(a.sede_pdf).localeCompare(String(b.sede_pdf)) || String(a.fecha).localeCompare(String(b.fecha)));
  $("rowsChecks").innerHTML = checks.map(r=>`<tr><td>${{esc(r.autogestion)}}</td><td>${{esc(r.sede_pdf)}}</td><td>${{esc(r.fecha)}}</td><td>${{esc(r.hora_inicial)}} - ${{esc(r.hora_final)}}</td><td>${{esc(r.tecnicos_nombres)}}</td><td class="num">${{fmt.format(num(r.tecnicos))}}</td><td class="num">${{fmt.format(num(r.horas_calculadas))}}</td><td class="num">${{fmt.format(num(r.autogestiones_pdf))}}</td><td>${{esc(checkPillars(r))}}</td><td>${{esc(r.archivo)}}</td></tr>`).join("");
  const checklistAlerts = DATA.alerts.map(r => ({{
    categoria: r.categoria || "Checklist",
    severidad: r.severidad || "Revisar",
    tipo: r.tipo,
    sede: r.sede_pdf,
    autogestion: r.autogestion,
    fecha: r.fecha,
    campo: r.tipo && r.tipo.toLowerCase().includes("hora") ? "Horario" : r.tipo && r.tipo.toLowerCase().includes("pilar") ? "Pilares" : "Checklist",
    valor_pdf: r.horas_efectivas_por_tecnico || r.autogestiones_pdf || "",
    valor_cronograma: "",
    diferencia: "",
    actividades_total_pdf: r.actividades_total_pdf || "",
    actividades_con_cobro_pdf: r.actividades_con_cobro_pdf || "",
    actividades_sin_cobro_pdf: r.actividades_sin_cobro_pdf || "",
    detalle: r.tecnicos_nombres ? `Técnicos: ${{r.tecnicos_nombres}}` : "",
    archivo: r.archivo
  }}));
  const siteAlerts = (DATA.site_alerts || []).map(r => {{
    const siteMetrics = sitePillarRowFor(r.sede);
    return {{
      ...r,
      autogestion:"",
      fecha:"",
      archivo:"",
      actividades_total_pdf: siteMetrics.actividades_total_pdf || "",
      actividades_con_cobro_pdf: siteMetrics.actividades_con_cobro_pdf || "",
      actividades_sin_cobro_pdf: siteMetrics.actividades_sin_cobro_pdf || "",
    }};
  }});
  const severityOrder = {{"Crítica": 0, "Revisar": 1, "Informativa": 2}};
  const alerts = [...siteAlerts, ...checklistAlerts].filter(includesQ).sort((a,b)=>(severityOrder[a.severidad]??9)-(severityOrder[b.severidad]??9) || String(a.categoria).localeCompare(String(b.categoria)) || String(a.sede).localeCompare(String(b.sede)));
  $("rowsAlerts").innerHTML = alerts.map(r=>`<tr><td>${{esc(r.categoria)}}</td><td>${{severityBadge(r.severidad)}}</td><td>${{esc(r.tipo)}}</td><td>${{esc(r.sede)}}</td><td>${{esc(r.autogestion)}}</td><td>${{esc(r.fecha)}}</td><td>${{esc(r.campo)}}</td><td class="num">${{esc(r.valor_pdf)}}</td><td class="num">${{esc(r.valor_cronograma)}}</td><td class="num">${{esc(r.diferencia)}}</td><td class="num">${{esc(r.actividades_total_pdf ?? "")}}</td><td class="num">${{esc(r.actividades_con_cobro_pdf ?? "")}}</td><td class="num">${{esc(r.actividades_sin_cobro_pdf ?? "")}}</td><td>${{esc(r.detalle)}}</td><td>${{esc(r.archivo)}}</td></tr>`).join("");
  const files = DATA.files.filter(includesQ).sort((a,b)=>String(a.carpeta).localeCompare(String(b.carpeta)) || String(a.archivo).localeCompare(String(b.archivo)));
  $("rowsFiles").innerHTML = files.map(r=>`<tr><td><a target="_blank" rel="noopener" href="${{esc(r.url)}}">${{esc(r.archivo)}}</a></td><td>${{esc(r.carpeta)}}</td><td>${{esc(r.sede_match)}}</td><td>${{esc(r.dias_archivo || "Sin fecha")}}</td><td>${{esc(r.fechas_pdf || "Sin fecha")}}</td><td class="num">${{fmt.format(num(r.checklists_pdf))}}</td><td class="num">${{fmt.format(num(r.horas_pdf))}}</td><td class="num">${{fmt.format(num(r.autogestiones_pdf))}}</td><td class="days">${{esc(r.pilares_pdf || "Sin lectura PDF")}}</td><td>${{esc(r.fuente_dias || "PDF")}}</td><td class="num">${{fmt.format(num(r.confianza_match))}}</td><td>${{esc(String(r.creado_sharepoint).slice(0,10))}}</td></tr>`).join("");
  const calendar = DATA.calendar_validation.filter(r=>includesQ(r) && (state.ues==="all" || r.ues===state.ues) && (state.status==="all" || r.estado===state.status) && (state.onlyAlerts==="all" || r.estado==="Incompleto")).sort((a,b)=>num(b.cantidad_falta_diligenciar)-num(a.cantidad_falta_diligenciar) || num(b.cantidad_digitada_falta_cargar)-num(a.cantidad_digitada_falta_cargar) || String(a.sede).localeCompare(String(b.sede)));
  $("rowsCalendar").innerHTML = calendar.map(r=>`<tr><td>${{esc(r.sede)}}</td><td>${{esc(r.ues)}}</td><td>${{esc(r.fecha_inicio)}}-${{esc(r.fecha_fin)}}<br><span class="small">${{esc(r.fuente_inicio_fin || "Cronog")}}</span></td><td>${{stack(r.cantidad_esperada,[`Esperados: ${{r.dias_esperados || "Sin rango"}}`,`Cuadritos General: ${{r.cuadritos_general || "Sin detalle"}}`, r.marcados_no_contados ? `Marcados no contados: ${{r.marcados_no_contados}}` : "", `Sábados sin visita: ${{r.sabados_sin_visita || "Ninguno"}}`,`Fuente: ${{r.fuente_esperado || "General MAYO"}}`])}}</td><td>${{stack(r.cantidad_digitada,[`Días: ${{r.dias_digitados || "Sin digitar"}}`, r.dias_digitados_fuera_esperado ? `Fuera esperado: ${{r.dias_digitados_fuera_esperado}}` : ""], r.dias_digitados_fuera_esperado ? "warn-text" : "")}}</td><td>${{stack(r.cantidad_sharepoint,[`Días: ${{r.dias_sharepoint || "Sin cargar"}}`, r.dias_sharepoint_fuera_esperado ? `Fuera esperado: ${{r.dias_sharepoint_fuera_esperado}}` : ""], r.dias_sharepoint_fuera_esperado ? "warn-text" : "")}}</td><td>${{stack(r.cantidad_falta_diligenciar,[r.dias_falta_diligenciar ? `Días: ${{r.dias_falta_diligenciar}}` : "Sin faltantes"], num(r.cantidad_falta_diligenciar) ? "bad-text" : "")}}</td><td>${{stack(r.cantidad_digitada_falta_cargar,[r.dias_digitados_falta_cargar ? `Días: ${{r.dias_digitados_falta_cargar}}` : "Sin faltantes"], num(r.cantidad_digitada_falta_cargar) ? "bad-text" : "")}}</td><td>${{badge(r.estado)}}<br><span class="small">${{esc(r.alerta || "Sin alertas")}}</span></td></tr>`).join("");
}}
function setTab(tab) {{ state.tab=tab; document.querySelectorAll(".tab").forEach(b=>b.classList.toggle("active",b.dataset.tab===tab)); {panel_names_json}.forEach(n=>{{ const panel=$("panel-"+n); if(panel) panel.classList.toggle("hidden",n!==tab); }}); }}
function render() {{ renderKpis(); bars(); renderTables(); if (INCLUDE_APP_VIEW) renderAppUploads(); renderDayCalc(); }}
function csv() {{
  const map={{sites:DATA.sites.filter(siteOk),pillars:DATA.site_rows.filter(pillarOk),checks:DATA.checklists.filter(includesQ),alerts:[...(DATA.site_alerts||[]), ...DATA.alerts].filter(includesQ),files:DATA.files.filter(includesQ),calendar:DATA.calendar_validation.filter(includesQ),app:appRows(),days:dayRows()}};
  const rows=map[state.tab]||[]; if(!rows.length) return; const headers=Object.keys(rows[0]);
  const text=[headers.join(",")].concat(rows.map(r=>headers.map(h=>`"${{String(typeof r[h]==='object'?JSON.stringify(r[h]):r[h]??'').replace(/"/g,'""')}}"`).join(","))).join("\\n");
  const blob=new Blob([text],{{type:"text/csv;charset=utf-8"}}); const url=URL.createObjectURL(blob); const a=document.createElement("a"); a.href=url; a.download=`dashboard_${{state.tab}}_mayo_2026.csv`; a.click(); URL.revokeObjectURL(url);
}}
function statusEndpointFrom(endpoint) {{
  const url = new URL(endpoint, window.location.href);
  url.pathname = "/status";
  url.search = "";
  url.hash = "";
  return url.toString();
}}

async function waitForCompletion(endpoint, status, btn) {{
  const statusUrl = statusEndpointFrom(endpoint);

  for (let intento = 1; intento <= 90; intento++) {{
    await new Promise(resolve => setTimeout(resolve, 5000));

    let data = {{}};

    try {{
      const res = await fetch(statusUrl, {{
        method: "GET",
        cache: "no-store"
      }});

      data = await res.json().catch(() => ({{ok:false}}));

      if (!res.ok || !data.ok) {{
        status.textContent = "Actualización iniciada. Consultando estado en GitHub Actions...";
        continue;
      }}
    }} catch (err) {{
      status.textContent = "Actualización iniciada. Reintentando consulta de estado...";
      continue;
    }}

    if (data.status === "queued") {{
      status.textContent = "🟡 Actualización en cola en GitHub Actions...";
      continue;
    }}

    if (data.status === "in_progress") {{
      status.textContent = "🔵 Actualización en proceso. Leyendo SharePoint y generando dashboard...";
      continue;
    }}

    if (data.status === "completed" && data.conclusion === "success") {{
      status.textContent = "✅ Actualización completada. Recargando dashboard...";
      setTimeout(() => {{
        location.href = location.pathname + (location.search ? location.search + "&" : "?") + "t=" + Date.now();
      }}, 1800);
      return;
    }}

    if (data.status === "completed") {{
      status.classList.add("bad-text");
      status.textContent = "❌ La actualización terminó con errores: " + (data.conclusion || "sin conclusión") + ". Revisa GitHub Actions.";
      btn.disabled = false;
      return;
    }}

    status.textContent = "Actualización iniciada. Estado actual: " + (data.status || "consultando...");
  }}

  status.classList.add("bad-text");
  status.textContent = "La actualización sigue tardando. Revisa GitHub Actions o recarga en unos minutos.";
  btn.disabled = false;
}}

async function refreshSharePoint() {{
  const btn=$("refreshBtn");
  const status=$("refreshStatus");
  const localHttp = location.protocol === "http:" && !GITHUB_REFRESH.repository;
  const endpoint = REFRESH_ENDPOINT || (location.protocol === "file:" ? "http://127.0.0.1:8787/refresh" : (localHttp ? "/refresh" : ""));
  btn.disabled = true;
  status.classList.remove("bad-text");
  if (!endpoint) {{
    if (GITHUB_REFRESH.repository) {{
      await triggerGitHubWorkflow(btn, status);
      return;
    }}
    status.classList.add("bad-text");
    status.textContent = "Este dashboard esta publicado en GitHub Pages. Para refrescarlo abre GitHub Actions y ejecuta el workflow 'Dashboard mantenimiento', o espera la actualizacion programada.";
    btn.disabled = false;
    return;
  }}
  status.textContent = "Actualizando desde SharePoint...";
  try {{
    const res = await fetch(endpoint, {{
      method:"POST",
      headers:{{"Content-Type":"application/json"}},
      body:JSON.stringify(REFRESH_CONFIG)
    }});
    const data = await res.json().catch(()=>({{ok:false,message:"Respuesta no valida del servidor local."}}));
    if (!res.ok || !data.ok) throw new Error(data.message || data.error || `HTTP ${{res.status}}`);

    status.textContent = "Solicitud enviada. Esperando a que GitHub Actions termine...";
    waitForCompletion(endpoint, status, btn);
    return;
  }} catch (err) {{
    status.classList.add("bad-text");
    status.textContent = "No se pudo iniciar la actualización. Detalle: " + err.message;
    btn.disabled = false;
  }}
}}
async function triggerGitHubWorkflow(btn, status) {{
  const repo = GITHUB_REFRESH.repository;
  const workflow = GITHUB_REFRESH.workflow;
  const token = prompt("Pega un token de GitHub con permiso Actions: Read and write para ejecutar el dashboard. No se guarda en el HTML.");
  if (!token) {{
    status.textContent = "Actualizacion cancelada. No se ingreso token de GitHub.";
    btn.disabled = false;
    return;
  }}
  status.textContent = "Solicitando ejecucion en GitHub Actions...";
  try {{
    const res = await fetch(`https://api.github.com/repos/${{repo}}/actions/workflows/${{encodeURIComponent(workflow)}}/dispatches`, {{
      method: "POST",
      headers: {{
        "Accept": "application/vnd.github+json",
        "Authorization": `Bearer ${{token.trim()}}`,
        "X-GitHub-Api-Version": "2022-11-28"
      }},
      body: JSON.stringify({{
        ref: GITHUB_REFRESH.branch || "main",
        inputs: {{ anio: GITHUB_REFRESH.anio, mes: GITHUB_REFRESH.mes }}
      }})
    }});
    if (!res.ok) {{
      const detail = await res.text();
      throw new Error(`GitHub HTTP ${{res.status}}: ${{detail.slice(0, 400)}}`);
    }}
    const actionsUrl = `https://github.com/${{repo}}/actions/workflows/${{workflow}}`;
    status.innerHTML = `Ejecucion solicitada en GitHub Actions. Cuando termine, descarga el artifact <strong>dashboard-mantenimiento</strong>. <a href="${{actionsUrl}}" target="_blank" rel="noopener">Abrir Actions</a>`;
    window.open(actionsUrl, "_blank", "noopener");
  }} catch (err) {{
    status.classList.add("bad-text");
    status.textContent = "No se pudo ejecutar GitHub Actions desde el HTML. Detalle: " + err.message;
  }} finally {{
    btn.disabled = false;
  }}
}}
$("search").addEventListener("input",e=>{{state.search=e.target.value;render();}});
$("status").addEventListener("change",e=>{{state.status=e.target.value;render();}});
$("ues").addEventListener("change",e=>{{state.ues=e.target.value;render();}});
$("onlyAlerts").addEventListener("change",e=>{{state.onlyAlerts=e.target.value;render();}});
$("csvBtn").addEventListener("click",csv);
$("refreshBtn").addEventListener("click",refreshSharePoint);
$("allDaysBtn").addEventListener("click",()=>{{state.selectedDays.clear();render();}});
$("clearDaysBtn").addEventListener("click",()=>{{state.selectedDays.clear();render();}});
document.querySelectorAll(".audit-status").forEach(b=>b.addEventListener("click",()=>{{state.appFilter=b.dataset.appFilter;document.querySelectorAll(".audit-status").forEach(x=>x.classList.toggle("active",x.dataset.appFilter===state.appFilter));render();}}));
document.querySelectorAll(".tab").forEach(b=>b.addEventListener("click",()=>setTab(b.dataset.tab)));
fillFilters(); render();
</script>
</body>
</html>
"""

OUTPUT.write_text(html, encoding="utf-8")
print(OUTPUT)
