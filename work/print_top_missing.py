import openpyxl

p = r"C:\Users\DELL\Documents\Codex\2026-06-03\files-mentioned-by-the-user-05\outputs\validacion_soportes_sharepoint_mayo_2026.xlsx"
wb = openpyxl.load_workbook(p, data_only=True)
ws = wb["Resumen por sede"]
headers = [c.value for c in ws[1]]
rows = []
for values in ws.iter_rows(min_row=2, values_only=True):
    row = dict(zip(headers, values))
    if row["cantidad_faltante"] and row["cantidad_faltante"] > 0:
        rows.append(row)
rows.sort(key=lambda r: r["cantidad_faltante"], reverse=True)
for row in rows[:12]:
    print(
        "{sede} - digitados {dig} - cubiertos {cub} - faltantes {fal} - dias faltantes: {dias}".format(
            sede=row["sede"],
            dig=row["cantidad_digitada"],
            cub=row["cantidad_cubierta"],
            fal=row["cantidad_faltante"],
            dias=row["dias_faltantes"],
        )
    )
