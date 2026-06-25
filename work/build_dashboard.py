import json
from pathlib import Path

import openpyxl


SOURCE = Path(r"C:\Users\DELL\Documents\Codex\2026-06-03\files-mentioned-by-the-user-05\outputs\validacion_soportes_sharepoint_mayo_2026.xlsx")
OUTPUT = Path(r"C:\Users\DELL\Documents\Codex\2026-06-03\files-mentioned-by-the-user-05\outputs\dashboard_soportes_mayo_2026.html")


def rows_from_sheet(wb, sheet_name):
    ws = wb[sheet_name]
    headers = [str(cell.value or "") for cell in ws[1]]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None and str(v).strip() for v in values):
            continue
        row = {}
        for header, value in zip(headers, values):
            row[header] = "" if value is None else value
        rows.append(row)
    return rows


wb = openpyxl.load_workbook(SOURCE, data_only=True)
data = {
    "overview": rows_from_sheet(wb, "Resumen"),
    "sites": rows_from_sheet(wb, "Resumen por sede"),
    "details": rows_from_sheet(wb, "Detalle por fecha"),
    "files": rows_from_sheet(wb, "Archivos SharePoint"),
    "unmatchedFiles": rows_from_sheet(wb, "Archivos sin match"),
}

payload = json.dumps(data, ensure_ascii=False)

html = f"""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Dashboard Soportes Mayo 2026</title>
  <style>
    :root {{
      --bg: #f6f7f9;
      --panel: #ffffff;
      --ink: #17202a;
      --muted: #657080;
      --line: #dbe1ea;
      --blue: #1769aa;
      --green: #1f8a5b;
      --red: #c83b3b;
      --amber: #b7791f;
      --gray: #edf1f6;
      --shadow: 0 10px 24px rgba(20, 31, 45, 0.08);
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      background: var(--bg);
      color: var(--ink);
      font-family: Arial, Helvetica, sans-serif;
      font-size: 14px;
      letter-spacing: 0;
    }}
    header {{
      background: #ffffff;
      border-bottom: 1px solid var(--line);
      padding: 18px 24px 14px;
      position: sticky;
      top: 0;
      z-index: 5;
    }}
    .title-row {{
      display: flex;
      align-items: flex-end;
      justify-content: space-between;
      gap: 18px;
      max-width: 1440px;
      margin: 0 auto;
    }}
    h1 {{
      margin: 0;
      font-size: 24px;
      line-height: 1.2;
      font-weight: 700;
    }}
    .subtitle {{
      margin-top: 4px;
      color: var(--muted);
      font-size: 13px;
    }}
    .toolbar {{
      display: grid;
      grid-template-columns: minmax(220px, 1.3fr) repeat(3, minmax(150px, 0.7fr));
      gap: 10px;
      max-width: 1440px;
      margin: 16px auto 0;
    }}
    input, select, button {{
      height: 36px;
      border: 1px solid var(--line);
      border-radius: 6px;
      background: #fff;
      color: var(--ink);
      padding: 0 10px;
      font: inherit;
      min-width: 0;
    }}
    button {{
      cursor: pointer;
      background: #113c5f;
      color: #fff;
      border-color: #113c5f;
      font-weight: 700;
    }}
    main {{
      max-width: 1440px;
      margin: 0 auto;
      padding: 18px 24px 36px;
    }}
    .kpis {{
      display: grid;
      grid-template-columns: repeat(6, minmax(130px, 1fr));
      gap: 12px;
      margin-bottom: 16px;
    }}
    .metric {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 13px 14px;
      box-shadow: var(--shadow);
      min-height: 86px;
    }}
    .metric span {{
      display: block;
      color: var(--muted);
      font-size: 12px;
      margin-bottom: 8px;
    }}
    .metric strong {{
      font-size: 26px;
      line-height: 1;
    }}
    .metric .note {{
      margin-top: 8px;
      font-size: 12px;
      color: var(--muted);
    }}
    .grid {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 16px;
      align-items: start;
    }}
    section {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 16px;
      box-shadow: var(--shadow);
      margin-bottom: 16px;
    }}
    h2 {{
      margin: 0 0 12px;
      font-size: 16px;
    }}
    .bar-list {{
      display: grid;
      gap: 9px;
    }}
    .bar-row {{
      display: grid;
      grid-template-columns: minmax(150px, 1fr) 3fr 46px;
      gap: 10px;
      align-items: center;
    }}
    .bar-label {{
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
      color: #243142;
    }}
    .bar-track {{
      height: 14px;
      background: var(--gray);
      border-radius: 4px;
      overflow: hidden;
    }}
    .bar-fill {{
      height: 100%;
      background: var(--red);
      border-radius: 4px;
    }}
    .bar-fill.green {{ background: var(--green); }}
    .bar-value {{
      text-align: right;
      color: var(--muted);
      font-variant-numeric: tabular-nums;
    }}
    .status-pill {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-width: 82px;
      padding: 4px 8px;
      border-radius: 999px;
      font-size: 12px;
      font-weight: 700;
    }}
    .complete {{ color: #0f6845; background: #ddf5e9; }}
    .incomplete {{ color: #9c2d2d; background: #fde2e2; }}
    .loaded {{ color: #0f6845; background: #ddf5e9; }}
    .missing {{ color: #9c2d2d; background: #fde2e2; }}
    .table-wrap {{
      overflow: auto;
      max-height: 560px;
      border: 1px solid var(--line);
      border-radius: 8px;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      min-width: 920px;
      background: #fff;
    }}
    th, td {{
      text-align: left;
      padding: 9px 10px;
      border-bottom: 1px solid var(--line);
      vertical-align: top;
    }}
    th {{
      position: sticky;
      top: 0;
      background: #eef3f8;
      z-index: 1;
      font-size: 12px;
      text-transform: uppercase;
      color: #465568;
    }}
    td.num, th.num {{ text-align: right; font-variant-numeric: tabular-nums; }}
    tr:hover td {{ background: #f8fbfe; }}
    .site-link {{
      border: 0;
      background: transparent;
      color: var(--blue);
      padding: 0;
      height: auto;
      font-weight: 700;
      text-align: left;
      cursor: pointer;
    }}
    .tabs {{
      display: flex;
      gap: 8px;
      margin: 2px 0 12px;
      flex-wrap: wrap;
    }}
    .tab {{
      height: 34px;
      border: 1px solid var(--line);
      background: #fff;
      color: var(--ink);
      border-radius: 6px;
      padding: 0 12px;
    }}
    .tab.active {{
      background: #113c5f;
      color: #fff;
      border-color: #113c5f;
    }}
    .hidden {{ display: none; }}
    .drawer {{
      border-top: 1px solid var(--line);
      margin-top: 12px;
      padding-top: 12px;
      color: var(--muted);
    }}
    .drawer strong {{ color: var(--ink); }}
    .file-link {{
      color: var(--blue);
      text-decoration: none;
    }}
    .file-link:hover {{ text-decoration: underline; }}
    @media (max-width: 1100px) {{
      .toolbar, .kpis, .grid {{ grid-template-columns: 1fr 1fr; }}
    }}
    @media (max-width: 720px) {{
      header {{ position: static; padding: 16px; }}
      main {{ padding: 16px; }}
      .title-row {{ align-items: flex-start; flex-direction: column; }}
      .toolbar, .kpis, .grid {{ grid-template-columns: 1fr; }}
      h1 {{ font-size: 20px; }}
      .metric strong {{ font-size: 22px; }}
      .bar-row {{ grid-template-columns: 1fr; gap: 5px; }}
    }}
  </style>
</head>
<body>
  <header>
    <div class="title-row">
      <div>
        <h1>Validación de soportes de visitas - Mayo 2026</h1>
        <div class="subtitle">Comparación entre fechas digitadas en Cronog Mayo 2026 y archivos cargados en SharePoint.</div>
      </div>
      <button id="exportCsv">Descargar vista CSV</button>
    </div>
    <div class="toolbar">
      <input id="search" type="search" placeholder="Buscar sede, UES, ciudad o archivo">
      <select id="statusFilter">
        <option value="all">Todos los estados</option>
        <option value="Incompleto">Incompletos</option>
        <option value="Completo">Completos</option>
      </select>
      <select id="uesFilter"><option value="all">Todas las UES</option></select>
      <select id="folderFilter"><option value="all">Todas las carpetas</option></select>
    </div>
  </header>

  <main>
    <div class="kpis">
      <div class="metric"><span>Archivos SharePoint</span><strong id="kpiFiles">0</strong><div class="note">PDFs encontrados</div></div>
      <div class="metric"><span>Sedes digitadas</span><strong id="kpiSites">0</strong><div class="note">Con fecha real</div></div>
      <div class="metric"><span>Checklists digitados</span><strong id="kpiTyped">0</strong><div class="note">Base de validación</div></div>
      <div class="metric"><span>Cubiertos</span><strong id="kpiCovered">0</strong><div class="note">Con soporte cargado</div></div>
      <div class="metric"><span>Faltantes</span><strong id="kpiMissing">0</strong><div class="note">Pendientes de soporte</div></div>
      <div class="metric"><span>Cobertura</span><strong id="kpiRate">0%</strong><div class="note">Cubiertos / digitados</div></div>
    </div>

    <div class="grid">
      <section>
        <h2>Mayores faltantes por sede</h2>
        <div id="missingBars" class="bar-list"></div>
      </section>
      <section>
        <h2>Cobertura por sede</h2>
        <div id="coverageBars" class="bar-list"></div>
      </section>
    </div>

    <section>
      <h2>Comparativo</h2>
      <div class="tabs">
        <button class="tab active" data-tab="sites">Resumen por sede</button>
        <button class="tab" data-tab="details">Detalle por fecha</button>
        <button class="tab" data-tab="files">Archivos cargados</button>
      </div>

      <div id="panel-sites" class="table-wrap">
        <table>
          <thead>
            <tr>
              <th>Sede</th>
              <th>UES</th>
              <th>Ubicación</th>
              <th class="num">Digitados</th>
              <th class="num">Cubiertos</th>
              <th class="num">Faltantes</th>
              <th>Días faltantes</th>
              <th>Estado</th>
            </tr>
          </thead>
          <tbody id="siteRows"></tbody>
        </table>
      </div>

      <div id="panel-details" class="table-wrap hidden">
        <table>
          <thead>
            <tr>
              <th>Sede</th>
              <th>Fecha digitada</th>
              <th>Estado</th>
              <th>Archivo soporte</th>
            </tr>
          </thead>
          <tbody id="detailRows"></tbody>
        </table>
      </div>

      <div id="panel-files" class="table-wrap hidden">
        <table>
          <thead>
            <tr>
              <th>Archivo</th>
              <th>Carpeta</th>
              <th>Sede asociada</th>
              <th>Días archivo</th>
              <th class="num">Confianza</th>
              <th>Fecha carga</th>
            </tr>
          </thead>
          <tbody id="fileRows"></tbody>
        </table>
      </div>

      <div id="drawer" class="drawer">Selecciona una sede para ver sus fechas y soportes asociados.</div>
    </section>
  </main>

  <script>
    const DATA = {payload};
    const state = {{ tab: "sites", selectedSite: "", search: "", status: "all", ues: "all", folder: "all" }};

    const fmt = new Intl.NumberFormat("es-CO");
    const $ = (id) => document.getElementById(id);
    const esc = (value) => String(value ?? "").replace(/[&<>"']/g, ch => ({{ "&": "&amp;", "<": "&lt;", ">": "&gt;", '"': "&quot;", "'": "&#39;" }}[ch]));
    const num = (value) => Number(value || 0);
    const norm = (value) => String(value ?? "").toLowerCase().normalize("NFD").replace(/[\\u0300-\\u036f]/g, "");

    function populateFilters() {{
      const ues = [...new Set(DATA.sites.map(r => r.ues).filter(Boolean))].sort();
      $("uesFilter").innerHTML += ues.map(v => `<option value="${{esc(v)}}">${{esc(v)}}</option>`).join("");
      const folders = [...new Set(DATA.files.map(r => r.carpeta).filter(Boolean))].sort();
      $("folderFilter").innerHTML += folders.map(v => `<option value="${{esc(v)}}">${{esc(v)}}</option>`).join("");
    }}

    function siteMatches(row) {{
      const q = norm(state.search);
      const text = norm([row.sede, row.ues, row.ubicacion, row.dias_faltantes].join(" "));
      const statusOk = state.status === "all" || row.estado === state.status;
      const uesOk = state.ues === "all" || row.ues === state.ues;
      const searchOk = !q || text.includes(q);
      return statusOk && uesOk && searchOk;
    }}

    function filteredSites() {{
      return DATA.sites.filter(siteMatches).sort((a, b) => num(b.cantidad_faltante) - num(a.cantidad_faltante) || String(a.sede).localeCompare(String(b.sede)));
    }}

    function filteredDetails() {{
      const siteSet = new Set(filteredSites().map(r => r.sede));
      const q = norm(state.search);
      return DATA.details.filter(row => {{
        const searchOk = !q || norm([row.sede, row.ues, row.ubicacion, row.archivo_soporte, row.estado].join(" ")).includes(q);
        return siteSet.has(row.sede) && searchOk;
      }}).sort((a, b) => String(a.sede).localeCompare(String(b.sede)) || num(a.dia) - num(b.dia));
    }}

    function filteredFiles() {{
      const q = norm(state.search);
      return DATA.files.filter(row => {{
        const searchOk = !q || norm([row.archivo, row.carpeta, row.sede_match, row.sede_archivo, row.dias_archivo].join(" ")).includes(q);
        const folderOk = state.folder === "all" || row.carpeta === state.folder;
        return searchOk && folderOk;
      }}).sort((a, b) => String(a.carpeta).localeCompare(String(b.carpeta)) || String(a.archivo).localeCompare(String(b.archivo)));
    }}

    function renderKpis(sites) {{
      const typed = sites.reduce((sum, r) => sum + num(r.cantidad_digitada), 0);
      const covered = sites.reduce((sum, r) => sum + num(r.cantidad_cubierta), 0);
      const missing = sites.reduce((sum, r) => sum + num(r.cantidad_faltante), 0);
      $("kpiFiles").textContent = fmt.format(filteredFiles().length);
      $("kpiSites").textContent = fmt.format(sites.length);
      $("kpiTyped").textContent = fmt.format(typed);
      $("kpiCovered").textContent = fmt.format(covered);
      $("kpiMissing").textContent = fmt.format(missing);
      $("kpiRate").textContent = typed ? Math.round((covered / typed) * 100) + "%" : "0%";
    }}

    function renderBars(sites) {{
      const maxMissing = Math.max(1, ...sites.map(r => num(r.cantidad_faltante)));
      $("missingBars").innerHTML = sites.filter(r => num(r.cantidad_faltante) > 0).slice(0, 12).map(row => `
        <div class="bar-row" title="${{esc(row.sede)}}">
          <div class="bar-label">${{esc(row.sede)}}</div>
          <div class="bar-track"><div class="bar-fill" style="width:${{Math.max(3, num(row.cantidad_faltante) / maxMissing * 100)}}%"></div></div>
          <div class="bar-value">${{fmt.format(num(row.cantidad_faltante))}}</div>
        </div>
      `).join("") || "<div class='subtitle'>No hay faltantes con los filtros actuales.</div>";

      const byCoverage = [...sites].filter(r => num(r.cantidad_digitada) > 0).sort((a, b) => (num(a.cantidad_cubierta) / num(a.cantidad_digitada)) - (num(b.cantidad_cubierta) / num(b.cantidad_digitada))).slice(0, 12);
      $("coverageBars").innerHTML = byCoverage.map(row => {{
        const rate = num(row.cantidad_digitada) ? Math.round(num(row.cantidad_cubierta) / num(row.cantidad_digitada) * 100) : 0;
        return `
          <div class="bar-row" title="${{esc(row.sede)}}">
            <div class="bar-label">${{esc(row.sede)}}</div>
            <div class="bar-track"><div class="bar-fill green" style="width:${{Math.max(3, rate)}}%"></div></div>
            <div class="bar-value">${{rate}}%</div>
          </div>
        `;
      }}).join("") || "<div class='subtitle'>No hay sedes para mostrar.</div>";
    }}

    function renderSiteRows(sites) {{
      $("siteRows").innerHTML = sites.map(row => `
        <tr>
          <td><button class="site-link" data-site="${{esc(row.sede)}}">${{esc(row.sede)}}</button></td>
          <td>${{esc(row.ues)}}</td>
          <td>${{esc(row.ubicacion)}}</td>
          <td class="num">${{fmt.format(num(row.cantidad_digitada))}}</td>
          <td class="num">${{fmt.format(num(row.cantidad_cubierta))}}</td>
          <td class="num">${{fmt.format(num(row.cantidad_faltante))}}</td>
          <td>${{esc(row.dias_faltantes || "Sin faltantes")}}</td>
          <td><span class="status-pill ${{row.estado === "Completo" ? "complete" : "incomplete"}}">${{esc(row.estado)}}</span></td>
        </tr>
      `).join("");
      document.querySelectorAll(".site-link").forEach(btn => btn.addEventListener("click", () => {{
        state.selectedSite = btn.dataset.site;
        renderDrawer();
      }}));
    }}

    function renderDetailRows(rows) {{
      $("detailRows").innerHTML = rows.map(row => `
        <tr>
          <td>${{esc(row.sede)}}</td>
          <td>${{esc(row.fecha_digitada)}}</td>
          <td><span class="status-pill ${{row.estado === "Cargado" ? "loaded" : "missing"}}">${{esc(row.estado)}}</span></td>
          <td>${{esc(row.archivo_soporte || "Sin archivo asociado")}}</td>
        </tr>
      `).join("");
    }}

    function renderFileRows(rows) {{
      $("fileRows").innerHTML = rows.map(row => `
        <tr>
          <td><a class="file-link" href="${{esc(row.url)}}">${{esc(row.archivo)}}</a></td>
          <td>${{esc(row.carpeta)}}</td>
          <td>${{esc(row.sede_match)}}</td>
          <td>${{esc(row.dias_archivo)}}</td>
          <td class="num">${{esc(row.confianza_match)}}</td>
          <td>${{esc(String(row.creado_sharepoint).slice(0, 10))}}</td>
        </tr>
      `).join("");
    }}

    function renderDrawer() {{
      if (!state.selectedSite) {{
        $("drawer").innerHTML = "Selecciona una sede para ver sus fechas y soportes asociados.";
        return;
      }}
      const site = DATA.sites.find(r => r.sede === state.selectedSite);
      const rows = DATA.details.filter(r => r.sede === state.selectedSite);
      const missing = rows.filter(r => r.estado !== "Cargado").map(r => r.dia).join(", ") || "Sin faltantes";
      const loaded = rows.filter(r => r.estado === "Cargado").map(r => `${{r.dia}}: ${{r.archivo_soporte}}`).join("<br>") || "Sin soportes asociados";
      $("drawer").innerHTML = `
        <strong>${{esc(state.selectedSite)}}</strong><br>
        Digitados: ${{fmt.format(num(site?.cantidad_digitada))}} | Cubiertos: ${{fmt.format(num(site?.cantidad_cubierta))}} | Faltantes: ${{fmt.format(num(site?.cantidad_faltante))}}<br>
        <strong>Días faltantes:</strong> ${{esc(missing)}}<br>
        <strong>Soportes asociados:</strong><br>${{loaded}}
      `;
    }}

    function switchTab(tab) {{
      state.tab = tab;
      document.querySelectorAll(".tab").forEach(btn => btn.classList.toggle("active", btn.dataset.tab === tab));
      ["sites", "details", "files"].forEach(name => $("panel-" + name).classList.toggle("hidden", name !== tab));
    }}

    function render() {{
      const sites = filteredSites();
      renderKpis(sites);
      renderBars(sites);
      renderSiteRows(sites);
      renderDetailRows(filteredDetails());
      renderFileRows(filteredFiles());
      renderDrawer();
    }}

    function downloadCsv() {{
      const rows = state.tab === "files" ? filteredFiles() : state.tab === "details" ? filteredDetails() : filteredSites();
      if (!rows.length) return;
      const headers = Object.keys(rows[0]);
      const csv = [headers.join(",")].concat(rows.map(row => headers.map(h => `"${{String(row[h] ?? "").replace(/"/g, '""')}}"`).join(","))).join("\\n");
      const blob = new Blob([csv], {{ type: "text/csv;charset=utf-8" }});
      const url = URL.createObjectURL(blob);
      const a = document.createElement("a");
      a.href = url;
      a.download = `vista_${{state.tab}}_soportes_mayo_2026.csv`;
      a.click();
      URL.revokeObjectURL(url);
    }}

    $("search").addEventListener("input", e => {{ state.search = e.target.value; render(); }});
    $("statusFilter").addEventListener("change", e => {{ state.status = e.target.value; render(); }});
    $("uesFilter").addEventListener("change", e => {{ state.ues = e.target.value; render(); }});
    $("folderFilter").addEventListener("change", e => {{ state.folder = e.target.value; render(); }});
    $("exportCsv").addEventListener("click", downloadCsv);
    document.querySelectorAll(".tab").forEach(btn => btn.addEventListener("click", () => switchTab(btn.dataset.tab)));

    populateFilters();
    render();
  </script>
</body>
</html>
"""

OUTPUT.write_text(html, encoding="utf-8")
print(OUTPUT)
